"""
Detailed Functional QA Report Generator
Divided into two main sections: Listing Page and Details Page.
"""

import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Palette
C = {
    "navy": "FF17233A", "navy_light": "FF1E3A5F", "white": "FFFFFFFF", "gray_bg": "FFF1F5F9",
    "notrun_bg": "FFF5F0FF", "notrun_fg": "FF5B21B6",
    "pos_bg": "FFD1FAE5", "pos_fg": "FF065F46",
    "neg_bg": "FFFEE2E2", "neg_fg": "FF991B1B",
}

def fill(h): return PatternFill("solid", fgColor=h)
def font(bold=False, color="FF000000", size=10): return Font(name="Calibri", size=size, bold=bold, color=color)
def align(h="left", wrap=True): return Alignment(horizontal=h, vertical="center", wrap_text=wrap)
def bdr(): return Border(left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin"))

TESTS = [
    # ============================================================================================================
    ("PAGE", "LISTING PAGE (Filters, Search, Virtualization, Pagination)"),
    # ============================================================================================================
    
    # ── SEARCH BAR ──────────────────────────────────────
    ("F-SRC-01", "Positive", "Search", "Input Debounce", 
     "1. Type 'test' in the search bar.\n2. Pause for 200ms.\n3. Type 'ing'.\n4. Pause for 600ms.", 
     "• Debounce resets on each keystroke.\n• API fires exactly once, ~500ms after typing stops.\n• Spinner is visible during the API flight.", "Not Run", "Critical"),
    ("F-SRC-02", "Positive", "Search", "Clear via Backspace", 
     "1. Highlight all text in the active search bar.\n2. Press Backspace to delete.", 
     "• Debounce triggers a new fetch.\n• API is called with an empty search parameter.\n• The list refreshes to the default state.", "Not Run", "High"),
    ("F-SRC-03", "Positive", "Search", "Clear via X Icon", 
     "1. Type text into the search bar.\n2. Click the 'X' inside the input field.", 
     "• Search text clears instantly.\n• API immediately triggers an empty fetch.", "Not Run", "Medium"),
    ("F-SRC-N1", "Negative", "Search", "XSS Injection", 
     "1. Click the search bar.\n2. Type exactly: `<script>alert('xss')</script>`", 
     "• React safely escapes HTML entities.\n• No malicious scripts are executed.\n• Returns an empty list state gracefully.", "Not Run", "Critical"),
    ("F-SRC-N2", "Negative", "Search", "Extreme Length", 
     "1. Copy a contiguous 10,000 character string.\n2. Paste it into the search input.", 
     "• Application does not freeze or crash.\n• Debounce fires safely.\n• API handles the length (truncates or returns 0 matches).", "Not Run", "High"),
    ("F-SRC-N3", "Negative", "Search", "SQL Injection", 
     "1. Type exactly: `' OR 1=1 --` into the search bar.", 
     "• Input is treated as a literal string (parameterized).\n• No database dump occurs.\n• Empty state safely returns.", "Not Run", "Critical"),

    # ── FILTER BUTTONS ──────────────────────────────────
    ("F-FLT-01", "Positive", "Filter", "Single Toggle", 
     "1. Click the 'Accepted' filter.\n2. Wait for the list to fetch.\n3. Click 'Accepted' a second time.", 
     "• First click applies active state UI and fetches `status=accepted`.\n• Second click turns the button off.\n• Reverts to the default route parameters.", "Not Run", "Critical"),
    ("F-FLT-02", "Positive", "Filter", "Age & Status Isolation", 
     "1. Navigate to the `/active` page.\n2. Click the `< 3 Days` filter.", 
     "• API sends `status=pending` AND `age=green`.\n• The age filter strictly respects the base active status without overriding it.", "Not Run", "Critical"),
    ("F-FLT-03", "Positive", "Filter", "Filter + Search Combo", 
     "1. Select the 'On Hold' filter.\n2. Wait for the list to load.\n3. Type 'Genoa' into the search bar.", 
     "• API request must include BOTH `status=On Hold` and the search string `Genoa`.\n• List reliably displays the exact intersection.", "Not Run", "High"),
    ("F-FLT-04", "Positive", "Filter", "Route Switching Reset", 
     "1. Apply a filter on the `/active` page.\n2. Click 'Completed List' in the navbar.", 
     "• Route changes successfully.\n• The previous filter state is completely destroyed.\n• The completed list fetches using its own default values.", "Not Run", "High"),
    ("F-FLT-N1", "Negative", "Filter", "Rapid Spam Clicking", 
     "1. Target any filter button.\n2. Click it rapidly 20 times within 2 seconds.", 
     "• App restricts fetch thrashing via abort controllers or local locking.\n• Only the single final state fetches.\n• UI avoids race conditions.", "Not Run", "High"),
    ("F-FLT-N2", "Negative", "Filter", "Network Disconnect", 
     "1. Disconnect network or mock offline mode.\n2. Click any filter button.", 
     "• Axios catches the failure.\n• ErrorBoundary or 'Network Error' snackbar safely appears.", "Not Run", "Medium"),

    # ── PAGINATION & LIST ───────────────────────────────
    ("F-PAG-01", "Positive", "Pagination", "Infinite Trigger", 
     "1. View a list with totalRecords > 50.\n2. Scroll slowly down towards the bottom of the items.", 
     "• The `loadMore` function triggers automatically.\n• A temporary bottom loader appears.\n• Next batch of records appends smoothly without jumping.", "Not Run", "Critical"),
    ("F-PAG-N1", "Negative", "Pagination", "Out of Bounds Limits", 
     "1. Mock or reach a state where you are on the very last page of records.\n2. Scroll down repeatedly.", 
     "• The client recognizes `page == totalPages`.\n• Scrolling triggers ZERO surplus API calls.\n• No loader is shown at the bottom.", "Not Run", "High"),
    ("F-PAG-N2", "Negative", "Pagination", "API Pagination Error", 
     "1. Scroll to trigger page 2 load.\n2. Force the API to return a 500 error on that page.", 
     "• A snackbar warns the user that fetching the next page failed.\n• Page 1 data remains securely intact and viewable.", "Not Run", "High"),

    # ── VIRTUALIZATION ──────────────────────────────────
    ("F-VRT-01", "Positive", "Virtualization", "Dynamic Row Heights", 
     "1. Slowly scroll through a dense list mingling 'Accepted' and 'Pending' documents.", 
     "• Logic assigns exact `100px` height strictly to Accepted targets.\n• `160px` assigned to Pending.\n• No DOM clipping/card overlapping.", "Not Run", "Critical"),
    ("F-VRT-02", "Positive", "Virtualization", "Threshold Load", 
     "1. Scroll deeply, landing exactly on the 5th item from the calculated bottom.", 
     "• `Threshold=5` explicitly forces the next data chunk automatically without scrolling the final remaining pixels.", "Not Run", "High"),
    ("F-VRT-N1", "Negative", "Virtualization", "Resize Throttling", 
     "1. Drag and violently resize Chrome browser width for 5 continuous seconds.", 
     "• React `AutoSizer` handles limits optimally.\n• Averts total 'ResizeObserver loop limit' UI freeze.", "Not Run", "Medium"),
    ("F-VRT-N2", "Negative", "Virtualization", "Skeleton Placement Bounds", 
     "1. Throttle network to 'Slow'.\n2. Scroll to load more records.", 
     "• Skeleton loader is safely positioned cleanly underneath all preceding indexes using specific list heights.", "Not Run", "Medium"),

    # ── LISTING NAV ─────────────────────────────────────
    ("F-NAV-01", "Positive", "Navigation", "Navbar Reset", 
     "1. Navigate to 'Active List'.\n2. Search/filter the data.\n3. Click 'Active List' in the navbar again.", 
     "• Forces the list back to page 1.\n• Clears the search input instantly.\n• Clears all applied filters.", "Not Run", "Medium"),
    ("F-NAV-02", "Positive", "Navigation", "Card Click Routing", 
     "1. Find any record card in a list.\n2. Click anywhere on the card body.", 
     "• The app dynamically navigates to `/:id`.\n• The details page opens correctly for that record.", "Not Run", "Critical"),

    # ============================================================================================================
    ("PAGE", "DETAILS PAGE (Corrections, Suggestions, Inline Edit, Forms & Dialogs)"),
    # ============================================================================================================

    # ── DETAILS NAV ─────────────────────────────────────
    ("F-NAV-N1", "Negative", "Navigation", "Invalid ID URL", 
     "1. Open the browser address bar.\n2. Manually navigate to `/invalid_execution_999`.", 
     "• API returns a 404 response.\n• Details component gracefully renders a 'Record not found' message.", "Not Run", "High"),

    # ── CORRECTIONS EXPANSION ───────────────────────────
    ("F-CXP-01", "Positive", "Corrections", "Default State Pending", 
     "1. Open a record containing mixed fields (Pending, Accepted, On Hold).", 
     "• Pending groups auto-expand immediately.\n• Accepted and On Hold groups auto-collapse by default.", "Not Run", "Critical"),
    ("F-CXP-02", "Positive", "Corrections", "Independent Toggles", 
     "1. Click the header of Group A.\n2. Click the header of Group B.", 
     "• Both groups expand/collapse independently.\n• Expanding one does not force the others to close (No accordion lock).", "Not Run", "Low"),
    ("F-CXP-N1", "Negative", "Corrections", "Malformed Status", 
     "1. Intercept network and set a field's `currentStatus` to `null`.", 
     "• Fallback logic kicks in.\n• App treats the field as 'pending' (auto-expands) rather than throwing a Javascript undefined crash.", "Not Run", "Medium"),

    # ── SUGGESTION ROW ──────────────────────────────────
    ("F-SGR-01", "Positive", "Suggestions", "Mutual Exclusivity", 
     "1. Select Suggestion A.\n2. Find the Custom Dropdown below it.\n3. Search and select a Custom Value.", 
     "• Activating the custom option completely deselects Suggestion A.\n• The radio button unchecks.", "Not Run", "Critical"),
    ("F-SGR-02", "Positive", "Suggestions", "Action Button Linkage", 
     "1. Have no suggestion selected.\n2. Click a suggestion radio button.", 
     "• The green 'Accept' button becomes completely ENABLED.\n• 'Send to L0' and 'Submit Draft' MUST immediately DISABLE.", "Not Run", "Critical"),
    ("F-SGR-03", "Positive", "Suggestions", "Read-Only Completed", 
     "1. Open a section labeled 'Accepted'.\n2. Try to click on the suggestions.", 
     "• Elements appear faded (opacity 0.5).\n• Pointer-events are disabled completely.\n• Selections cannot be manipulated.", "Not Run", "High"),
    ("F-SGR-N1", "Negative", "Suggestions", "Double Selection Exploit", 
     "1. Focus Suggestion A.\n2. Rapidly switch state using Keyboard Tab and Spacebar.", 
     "• Singular selection rules hold firm.\n• It is impossible to check two radio buttons inside the exact same group.", "Not Run", "High"),
    ("F-SGR-N2", "Negative", "Suggestions", "Missing Data Keys", 
     "1. Inject a suggestion missing its primary key payload.", 
     "• Row render does not crash.\n• Empty attributes fallback safely to '-' or a blank visual state.", "Not Run", "Medium"),

    # ── VM CPU EDITING ──────────────────────────────────
    ("F-CPU-01", "Positive", "Inline Edit", "Initiation & Save", 
     "1. Click the Pencil icon next to CPU(s).\n2. Type a new number.\n3. Press Enter.", 
     "• Input box transforms back into text.\n• The new locally edited value relies in `editedSuggestions` state.", "Not Run", "High"),
    ("F-CPU-02", "Positive", "Inline Edit", "Cancel via Esc", 
     "1. Click the Pencil icon.\n2. Mutate the field number.\n3. Press the Escape key.", 
     "• The input window instantly closes.\n• The field immediately reverts to the exact original fetched suggestion.", "Not Run", "High"),
    ("F-CPU-03", "Positive", "Inline Edit", "Payload Injection", 
     "1. Edit the CPU(s) value to `77`.\n2. Accept the record.\n3. Examine Network tab.", 
     "• The `PUT /approve` MUST prioritize and transmit the manually typed integer `77`.", "Not Run", "Critical"),
    ("F-CPU-N1", "Negative", "Inline Edit", "Float Block / NaN", 
     "1. Click Pencil icon.\n2. Type the string 'abc'.\n3. Press Enter.", 
     "• The UI securely reverts the edit or loudly blocks the save via a 'numbers only' constraint.", "Not Run", "High"),
    ("F-CPU-N2", "Negative", "Inline Edit", "Negative Int Limit", 
     "1. Click Pencil icon.\n2. Type the value '-99'.\n3. Press Enter.", 
     "• Values falling beneath 0 are constrained and blocked from persisting to local state.", "Not Run", "High"),

    # ── CUSTOM VALUE DROPDOWN ───────────────────────────
    ("F-CVD-01", "Positive", "Custom Dropdown", "Metadata Fetch Flow", 
     "1. Click the Autocomplete dropdown.\n2. Select an unlisted value.", 
     "• The API fetches the new value's metadata.\n• A brand new dynamic SuggestionRow renders on-screen correctly.", "Not Run", "High"),
    ("F-CVD-02", "Positive", "Custom Dropdown", "Clear Selection", 
     "1. With a custom value loaded, click the 'X' button inside the input text field.", 
     "• The Custom Value SuggestionRow disappears.\n• Group selection is cleared completely out.", "Not Run", "High"),
    ("F-CVD-N1", "Negative", "Custom Dropdown", "Fetch Timeout", 
     "1. Select a Custom Value.\n2. The system triggers a mock 500 error or Network Timeout.", 
     "• Loading spinner halts.\n• An explicit Red error snackbar surfaces.\n• The Dropdown securely averts a crash by reverting empty.", "Not Run", "Critical"),
    ("F-CVD-N2", "Negative", "Custom Dropdown", "Garbage String Submission", 
     "1. Type complete gibberish into dropdown.\n2. Press Enter.", 
     "• Given the value cannot exist in the masterlist API, the UI rejects or strips the selection securely.", "Not Run", "Medium"),

    # ── L0 & ACCEPT FLOWS ───────────────────────────────
    ("F-FLW-01", "Positive", "L0 Flow", "Submission", 
     "1. Click the Red 'Send to L0' button.\n2. Click 'Yes, Confirm' in the prompt.", 
     "• Fires `PUT /reject-record`.\n• On 200 OK, Dialog clears, app data fetches, Green success alert mounts.", "Not Run", "Critical"),
    ("F-FLW-02", "Positive", "Accept Flow", "VM History Override", 
     "1. Find a VM record displaying prior `cpu(s)` history.\n2. Select a suggestion.\n3. Click Accept.", 
     "• The code logic silences the raw suggestion value and strictly submits the `history.changes` CPU count into the payload.", "Not Run", "Critical"),
    ("F-FLW-N1", "Negative", "Submit Flows", "500 Server Error", 
     "1. Initiate an Accept submission.\n2. The Server intercepts with a 500 Internal error.", 
     "• Dialog doesn't break. Remains open safely with unlocked buttons.\n• A concise Red snackbar alerts the user to the server fault.", "Not Run", "Critical"),
    ("F-FLW-N2", "Negative", "Submit Flows", "Mid-flight Cancel Lockout", 
     "1. Click Submit (mock a 2-second API delay).\n2. Attempt instantly clicking 'Cancel' or hitting 'Esc'.", 
     "• Hard lockout engages.\n• Dialog explicitly refuses to close during flight to strictly prevent duplicate POST/PUT requests.", "Not Run", "High"),

    # ── DRAFT FLOW ──────────────────────────────────────
    ("F-DRF-01", "Positive", "Draft Flow", "Dynamic Field Loading", 
     "1. Click the Yellow 'Submit Draft Record' button.", 
     "• Dialog spawns immediately.\n• Fetches `/draft-records/fields` dynamically.\n• Shows spinner while building UI.", "Not Run", "Critical"),
    ("F-DRF-02", "Positive", "Draft Flow", "History Pre-fill", 
     "1. Open Draft dialog for a field bearing known historical modifications.", 
     "• Dynamic inputs are NOT completely empty.\n• They pre-fill mapped to strings from `getHistoryChangesForField` local logic.", "Not Run", "High"),
    ("F-DRF-03", "Positive", "Draft Flow", "Form Validation Unlock", 
     "1. Observe Submit button is disabled initially.\n2. Methodically fill all required input text fields.", 
     "• Re-evaluates form state instantly.\n• Submit button enables exclusively when zero fields remain blank.", "Not Run", "High"),
    ("F-DRF-N1", "Negative", "Draft Flow", "Float/Decimal Rejection", 
     "1. Locate an input tied to `datatype: integer`.\n2. Type '14.5'.\n3. Press Submit.", 
     "• Halts hard at the local browser boundary.\n• Specific snackbar identifies float-rejection blocking.", "Not Run", "Critical"),
    ("F-DRF-N2", "Negative", "Draft Flow", "Negative Number Reject", 
     "1. On an `integer` input form, type '-5'.\n2. Press Submit.", 
     "• Constraint `min=0` enforces boundary.\n• Local logic refuses negative integer mutation and halts API.", "Not Run", "High"),
    ("F-DRF-N3", "Negative", "Draft Flow", "Whitespace Bypass", 
     "1. Empty a field and fill with exactly 5 spaces '     '.\n2. Evaluate validation.", 
     "• Field is functionally blank.\n• Local check blocks 'Submit' button by running `.trim()` validations.", "Not Run", "Medium"),
    ("F-DRF-N4", "Negative", "Draft Flow", "Missing Field Setup", 
     "1. Manipulate response giving a completely empty array of `fields[]` back.", 
     "• UI averts fatal crash.\n• Reports gracefully to user that draft attributes couldn't render.", "Not Run", "Low"),

    # ── FEEDBACK & ALERTS ────────────────────────────────
    ("F-FDB-01", "Positive", "Feedback", "Snackbar Timing", 
     "1. Cause a notification alert.\n2. Observe behavior and wait.", 
     "• Disappears fully unsupervised after exactly 3000ms.", "Not Run", "Medium"),
    ("F-FDB-N1", "Negative", "Feedback", "Snackbar Stacking", 
     "1. Trigger an internal error alert.\n2. While alert is visible, trigger an accept success response instantly.", 
     "• System maintains clear stack hierarchy.\n• UI overlay text doesn't corrupt or improperly bind to preceding colors.", "Not Run", "Low"),
]

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Categorized Functional QA"
ws.sheet_view.showGridLines = False

# Header
ws.merge_cells("A1:H1")
header = ws.cell(row=1, column=1, value="PAGE CATEGORIZED EXHAUSTIVE QA: All Interactive Flows + Pos/Neg Testing")
header.fill = fill(C["navy"]); header.font = font(bold=True, size=15, color=C["white"])
header.alignment = align("center")
ws.row_dimensions[1].height = 35

cols = ["TC ID", "Test Type", "Functional Area", "Component", "Steps", "Expected Behavior", "Status", "Priority"]
widths = [11, 14, 20, 24, 45, 55, 12, 12]

# Headers
for i, h in enumerate(cols, 1):
    cell = ws.cell(row=2, column=i, value=h)
    cell.fill = fill(C["navy"]); cell.font = font(bold=True, color=C["white"]); cell.alignment = align("center")
    ws.column_dimensions[get_column_letter(i)].width = widths[i-1]

row_idx = 3

for item in TESTS:
    # If this is a Section Banner
    if item[0] == "PAGE":
        ws.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=8)
        banner = ws.cell(row=row_idx, column=1, value=item[1])
        banner.fill = fill(C["navy_light"])
        banner.font = font(bold=True, size=13, color=C["white"])
        banner.alignment = align("center")
        ws.row_dimensions[row_idx].height = 30
        row_idx += 1
        continue
        
    (tcid, ttype, area, comp, steps, expt, stat, pri) = item
    
    bg = C["white"] if row_idx % 2 == 0 else C["gray_bg"]
    
    ws.cell(row=row_idx, column=1, value=tcid).border = bdr()
    ws.cell(row=row_idx, column=1).fill = fill(bg)
    
    type_cell = ws.cell(row=row_idx, column=2, value=ttype)
    type_cell.border = bdr(); type_cell.alignment = align("center")
    if ttype == "Positive":
        type_cell.fill = fill(C["pos_bg"]); type_cell.font = font(color=C["pos_fg"], bold=True)
    else:
        type_cell.fill = fill(C["neg_bg"]); type_cell.font = font(color=C["neg_fg"], bold=True)
        
    for c, val in enumerate([area, comp, steps, expt], 3):
        cell = ws.cell(row=row_idx, column=c, value=val)
        cell.fill = fill(bg); cell.alignment = align(); cell.font = font(); cell.border = bdr()
    
    s_cell = ws.cell(row=row_idx, column=7, value=stat)
    s_cell.fill = fill(C["notrun_bg"]); s_cell.font = font(bold=True, color=C["notrun_fg"]); s_cell.alignment = align("center"); s_cell.border = bdr()
    
    p_cell = ws.cell(row=row_idx, column=8, value=pri)
    p_cell.fill = fill(bg); p_cell.alignment = align("center"); p_cell.border = bdr()
    p_color = {"Critical": "FF991B1B", "High": "FF9A3412", "Medium": "FF854D0E", "Low": "FF166534"}.get(pri, "FF000000")
    p_cell.font = font(bold=True, color=p_color)

    row_idx += 1

OUTPUT_PATH = r"d:\DH\Data-Hygiene\QA_Categorized_Functional_Report.xlsx"
wb.save(OUTPUT_PATH)
print(f"Report saved to: {OUTPUT_PATH}")

"""
QA Report Generator — AMD Data Hygiene Frontend
Generates a fully-formatted Excel workbook covering all UI / frontend test cases.
"""

import openpyxl
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side, GradientFill
)
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, Reference
from openpyxl.chart.series import DataPoint
from datetime import date

# ─────────────────────────────────────────────────────────────────────────────
# PALETTE
# ─────────────────────────────────────────────────────────────────────────────
C = {
    "navy":       "FF17233A",
    "navy_light": "FF1E3A5F",
    "blue":       "FF1E6FD9",
    "blue_light": "FFD6E8FF",
    "white":      "FFFFFFFF",
    "off_white":  "FFF8FAFD",
    "gray_bg":    "FFF1F5F9",
    "gray_text":  "FF64748B",
    "header_row": "FFE2ECF9",

    # Status colours
    "pass_bg":    "FFD1FAE5",
    "pass_fg":    "FF065F46",
    "fail_bg":    "FFFEE2E2",
    "fail_fg":    "FF991B1B",
    "blocked_bg": "FFFEF3C7",
    "blocked_fg": "FF92400E",
    "na_bg":      "FFF1F5F9",
    "na_fg":      "FF334155",
    "notrun_bg":  "FFF5F0FF",
    "notrun_fg":  "FF5B21B6",

    # Priority
    "p_crit":     "FFEF4444",
    "p_high":     "FFf97316",
    "p_med":      "FFEab308",
    "p_low":      "FF22C55E",
}

def fill(hex_color): return PatternFill("solid", fgColor=hex_color)
def font(name="Calibri", size=11, bold=False, color="FF000000", italic=False):
    return Font(name=name, size=size, bold=bold, color=color, italic=italic)
def align(h="left", v="center", wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)
def border_all():
    s = Side(style="thin", color="FFD1D5DB")
    return Border(left=s, right=s, top=s, bottom=s)
def border_outer():
    s = Side(style="medium", color="FF94A3B8")
    return Border(left=s, right=s, top=s, bottom=s)

STATUS_STYLE = {
    "Pass":     (C["pass_bg"],    C["pass_fg"]),
    "Fail":     (C["fail_bg"],    C["fail_fg"]),
    "Blocked":  (C["blocked_bg"], C["blocked_fg"]),
    "N/A":      (C["na_bg"],      C["na_fg"]),
    "Not Run":  (C["notrun_bg"],  C["notrun_fg"]),
}

PRIORITY_COLOR = {
    "Critical": C["p_crit"],
    "High":     C["p_high"],
    "Medium":   C["p_med"],
    "Low":      C["p_low"],
}

def apply_status(cell, status):
    if status in STATUS_STYLE:
        bg, fg = STATUS_STYLE[status]
        cell.fill = fill(bg)
        cell.font = font(bold=True, color=fg, size=10)
    cell.alignment = align("center")
    cell.border = border_all()

def apply_priority(cell, pri):
    color = PRIORITY_COLOR.get(pri, C["na_fg"])
    cell.fill = fill("FFFFFFFF")
    cell.font = font(bold=True, color=color, size=10)
    cell.alignment = align("center")
    cell.border = border_all()

def write_cell(ws, row, col, value, bg=None, fg="FF000000", bold=False,
               size=11, h_align="left", wrap=False, italic=False, border=True):
    cell = ws.cell(row=row, column=col, value=value)
    if bg:  cell.fill = fill(bg)
    cell.font = font(size=size, bold=bold, color=fg, italic=italic)
    cell.alignment = align(h_align, wrap=wrap)
    if border: cell.border = border_all()
    return cell

def section_header(ws, row, col_count, title, bg=C["navy"], fg=C["white"]):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=col_count)
    cell = ws.cell(row=row, column=1, value=title)
    cell.fill = fill(bg)
    cell.font = font(size=12, bold=True, color=fg)
    cell.alignment = align("center")
    cell.border = border_outer()
    return cell

def col_header(ws, row, col, value):
    cell = ws.cell(row=row, column=col, value=value)
    cell.fill = fill(C["header_row"])
    cell.font = font(size=10, bold=True, color=C["navy"])
    cell.alignment = align("center")
    cell.border = border_all()
    return cell

# ─────────────────────────────────────────────────────────────────────────────
# TEST CASE DATA
# ─────────────────────────────────────────────────────────────────────────────

# Columns: TC_ID, Module, Component, Test Case Description, Pre-conditions,
#          Steps, Expected Result, Actual Result, Status, Priority, Remarks

TESTS = [

    # ── 1. NAVIGATION / ROUTING ──────────────────────────────────────────────
    ("TC-NAV-01","Navigation","Navbar",
     "Logo click navigates to home",
     "App is loaded",
     "1. Click the 'AMD_DH' logo in the navbar",
     "Route changes to '/' and Landing (All) list is displayed",
     "","Not Run","High",""),

    ("TC-NAV-02","Navigation","Navbar",
     "Active List nav link navigates to /active",
     "App is loaded on any page",
     "1. Click 'Active List' in navbar",
     "Route changes to '/active'; page title shows 'My Active List'",
     "","Not Run","High",""),

    ("TC-NAV-03","Navigation","Navbar",
     "Completed List nav link navigates to /completed",
     "App loaded",
     "1. Click 'Completed List' in navbar",
     "Route changes to '/completed'; page title shows 'My Completed List'",
     "","Not Run","High",""),

    ("TC-NAV-04","Navigation","Navbar",
     "On Hold nav link navigates to /on-hold",
     "App loaded",
     "1. Click 'On Hold' in navbar",
     "Route changes to '/on-hold'; page title shows 'On Hold Records'",
     "","Not Run","High",""),

    ("TC-NAV-05","Navigation","Navbar",
     "Active NavLink shows underline when on /active route",
     "User is on /active",
     "1. Observe the navbar links",
     "'Active List' link has bottomBorder '2px solid #fff'; others are dimmed (opacity 0.7)",
     "","Not Run","Medium",""),

    ("TC-NAV-06","Navigation","Navbar",
     "Navbar remains fixed on scroll",
     "User is on list page with many records",
     "1. Scroll down beyond the viewport 2. Observe navbar position",
     "Navbar stays pinned to the top (position: fixed)",
     "","Not Run","Medium",""),

    ("TC-NAV-07","Navigation","Routing",
     "Direct URL navigation to /:id opens the DetailsPage",
     "A valid Execution ID is known",
     "1. Navigate directly to /<valid-id> in the browser URL bar",
     "DetailsPage loads with the correct Execution Information and Corrections Table",
     "","Not Run","High",""),

    ("TC-NAV-08","Navigation","Routing",
     "Unknown route falls back gracefully",
     "App is loaded",
     "1. Navigate to an invalid route, e.g. /xyz-unknown-page",
     "App does not crash; either shows empty state or redirects to home",
     "","Not Run","Low",""),

    # ── 2. LANDING PAGE (All Records) ────────────────────────────────────────
    ("TC-LST-01","List Pages","RecordsListPage (landing)",
     "Landing page renders with correct title",
     "App loaded at '/'",
     "1. Open '/'",
     "Header displays 'Data Hygiene Dashboard'",
     "","Not Run","High",""),

    ("TC-LST-02","List Pages","RecordsListPage (landing)",
     "Status filter buttons shown on landing page",
     "App loaded at '/'",
     "1. Open '/'",
     "Filter buttons for 'Pending', 'Accepted', 'Rejected', 'On Hold' are visible",
     "","Not Run","High",""),

    ("TC-LST-03","List Pages","RecordsListPage (landing)",
     "Clicking a status filter fetches filtered records",
     "Landing page loaded with records",
     "1. Click 'Accepted' filter button",
     "Only accepted records are shown; count label updates",
     "","Not Run","High",""),

    ("TC-LST-04","List Pages","RecordsListPage (landing)",
     "Clicking active filter again deactivates it (toggle)",
     "Filter 'Accepted' is active",
     "1. Click 'Accepted' again",
     "Filter deactivates; all records are shown again",
     "","Not Run","High",""),

    # ── 3. ACTIVE LIST ───────────────────────────────────────────────────────
    ("TC-LST-05","List Pages","RecordsListPage (active)",
     "Active list page shows age filter buttons",
     "User navigates to /active",
     "1. Open '/active'",
     "'< 3 Days', '3 - 6 Days', '> 6 Days' filter buttons are visible; no status filters",
     "","Not Run","High",""),

    ("TC-LST-06","List Pages","RecordsListPage (active)",
     "Age filter '<3 Days' sends age=green to API",
     "/active loaded",
     "1. Click '< 3 Days' filter button; 2. Observe network request",
     "API called with age=green param; cards with green border appear",
     "","Not Run","High",""),

    ("TC-LST-07","List Pages","RecordsListPage (active)",
     "Age filter '3-6 Days' sends age=yellow to API",
     "/active loaded",
     "1. Click '3 - 6 Days' filter button",
     "API called with age=yellow; appropriate records shown",
     "","Not Run","High",""),

    ("TC-LST-08","List Pages","RecordsListPage (active)",
     "Age filter '>6 Days' sends age=red to API",
     "/active loaded",
     "1. Click '> 6 Days' filter button",
     "API called with age=red; appropriate records shown",
     "","Not Run","High",""),

    ("TC-LST-09","List Pages","RecordsListPage (active)",
     "Active page default status is 'pending'",
     "/active loaded without filter",
     "1. Open /active; 2. Observe network request",
     "API is called with status=pending by default",
     "","Not Run","High",""),

    # ── 4. COMPLETED LIST ────────────────────────────────────────────────────
    ("TC-LST-10","List Pages","RecordsListPage (completed)",
     "Completed list shows Accepted and Rejected filter buttons only",
     "User on /completed",
     "1. Open '/completed'",
     "Only 'Accepted' and 'Rejected' filters shown (Pending and On Hold are absent)",
     "","Not Run","High",""),

    ("TC-LST-11","List Pages","RecordsListPage (completed)",
     "Completed list default fetches accepted+rejected records",
     "User on /completed",
     "1. Open '/completed'; 2. Check network request",
     "API called with status=accepted,rejected by default",
     "","Not Run","Critical",""),

    # ── 5. ON-HOLD LIST ──────────────────────────────────────────────────────
    ("TC-LST-12","List Pages","RecordsListPage (onhold)",
     "On Hold list uses age filters",
     "User on /on-hold",
     "1. Open '/on-hold'",
     "Age filter buttons (< 3 Days, 3–6 Days, > 6 Days) are displayed",
     "","Not Run","High",""),

    ("TC-LST-13","List Pages","RecordsListPage (onhold)",
     "On Hold list default status is 'On Hold'",
     "User on /on-hold; no age filter selected",
     "1. Open /on-hold; 2. Inspect API request",
     "API called with status=On Hold",
     "","Not Run","High",""),

    # ── 6. LIST HEADER COMPONENT ─────────────────────────────────────────────
    ("TC-HDR-01","ListHeader","Search Bar",
     "Search input is visible and focusable",
     "Any list page loaded",
     "1. Click inside the search field",
     "Search field gains focus; cursor appears inside",
     "","Not Run","High",""),

    ("TC-HDR-02","ListHeader","Search Bar",
     "Search debounce: API called 500ms after typing stops",
     "List page loaded",
     "1. Type 'AMD' in search box quickly; 2. Wait 600ms",
     "Spinner appears in search field during type; API called once after 500ms with search=AMD",
     "","Not Run","High",""),

    ("TC-HDR-03","ListHeader","Search Bar",
     "Clearing search re-fetches all records",
     "Search term 'AMD' active",
     "1. Clear the search field (backspace to empty)",
     "API is called without search param; all records re-appear",
     "","Not Run","High",""),

    ("TC-HDR-04","ListHeader","Search Bar",
     "CircularProgress spinner shown while search loads",
     "List page loaded",
     "1. Start typing in search field before previous request completes",
     "A small circular progress icon appears at the right side of the search field",
     "","Not Run","Medium",""),

    ("TC-HDR-05","ListHeader","Search Bar",
     "SearchIcon always visible in search field start adornment",
     "Any list page",
     "1. Observe the search field at rest (no typing)",
     "Magnifier icon (#94a3b8 color) is shown inside the left side of the search input",
     "","Not Run","Low",""),

    ("TC-HDR-06","ListHeader","Filter Buttons",
     "Active filter button shows raised/highlighted state",
     "A filter is clicked",
     "1. Click any filter button",
     "Button background changes to activeBg color, font weight becomes 700, translateY(-2px) lift visible",
     "","Not Run","High",""),

    ("TC-HDR-07","ListHeader","Filter Buttons",
     "Filter button hover shows lift animation",
     "List page loaded",
     "1. Hover over a filter button without clicking",
     "Button lifts (translateY -2px) and shows shadow on hover",
     "","Not Run","Medium",""),

    ("TC-HDR-08","ListHeader","Filter Buttons",
     "Age filter dot changes color when active",
     "/active page loaded",
     "1. Click '< 3 Days' filter",
     "The colored dot inside the filter button turns white (#fff) when filter is active",
     "","Not Run","Low",""),

    # ── 7. RECORD LIST (Virtualization) ──────────────────────────────────────
    ("TC-VRL-01","RecordList","Virtualization",
     "Initial load spinner shown when records list is empty + loading",
     "List page opened fresh",
     "1. Open any list page immediately",
     "Full-screen Loader spinner shown while first page fetch is in-flight",
     "","Not Run","High",""),

    ("TC-VRL-02","RecordList","Virtualization",
     "Virtualized list renders only visible rows",
     "List with 50+ records loaded",
     "1. Open /active with many records; 2. Inspect DOM",
     "Only visible rows are rendered in DOM (react-virtualized behavior)",
     "","Not Run","Medium",""),

    ("TC-VRL-03","RecordList","Infinite Scroll",
     "Scrolling near bottom triggers loadMore (infinite scroll)",
     "First page (50 records) loaded; more pages available",
     "1. Scroll to the bottom of the list",
     "Additional records are fetched and appended (threshold = 5 rows from bottom)",
     "","Not Run","Critical",""),

    ("TC-VRL-04","RecordList","Infinite Scroll",
     "Bottom loader shown while next page is loading",
     "Scrolled to bottom with more pages available",
     "1. Scroll to bottom; 2. Observe while new rows load",
     "Loader spinner appears below the last visible card",
     "","Not Run","Medium",""),

    ("TC-VRL-05","RecordList","Count Label",
     "Total records count label displayed",
     "List loaded",
     "1. Observe count label above the list",
     "Text 'Total records: <N>' is shown when showCount is true",
     "","Not Run","Medium",""),

    ("TC-VRL-06","RecordList","Empty State",
     "Empty state message shown when no records match filter",
     "Filter applied with 0 results",
     "1. Apply a filter that returns 0 records",
     "'No records match the selected filter.' text displayed",
     "","Not Run","High",""),

    ("TC-VRL-07","RecordList","Empty State",
     "Default empty message when no records exist in category",
     "Navigating to a list with 0 total records",
     "1. Open a list category with no data",
     "'No records found in this category.' text displayed",
     "","Not Run","Medium",""),

    ("TC-VRL-08","RecordList","Row Height",
     "Accepted/Approved cards have height 100; other cards have height 160",
     "List with mixed status records",
     "1. Scroll through a list containing accepted and pending records",
     "Accepted cards are more compact visually; pending cards are taller (include Inconsistent Fields section)",
     "","Not Run","Medium",""),

    # ── 8. RECORD CARD ───────────────────────────────────────────────────────
    ("TC-RCD-01","RecordCard","Display",
     "Card shows ExecutionId, BenchmarkCategory, BenchmarkType, Status",
     "List page loaded with records",
     "1. Observe any record card",
     "All four labels and their values are displayed correctly",
     "","Not Run","Critical",""),

    ("TC-RCD-02","RecordCard","Status Colors",
     "Pending card has orange border and background",
     "List has a pending record",
     "1. Find a pending record",
     "Card has #ffa000 border, #fff8e1 background, orange status text",
     "","Not Run","High",""),

    ("TC-RCD-03","RecordCard","Status Colors",
     "Accepted card has green border and background",
     "List has an accepted record",
     "1. Find an accepted record",
     "Card has #43a047 border, #e8f5e9 background, green status text",
     "","Not Run","High",""),

    ("TC-RCD-04","RecordCard","Status Colors",
     "Rejected card has red border and background",
     "List has a rejected record",
     "1. Find a rejected record",
     "Card has #e53935 border, #ffebee background, red status text",
     "","Not Run","High",""),

    ("TC-RCD-05","RecordCard","Status Colors",
     "On Hold card has purple border and background",
     "List has an on-hold record",
     "1. Find an on-hold record",
     "Card has #7c3aed border, #ede9fe background, purple status text",
     "","Not Run","High",""),

    ("TC-RCD-06","RecordCard","Age Colors",
     "Age color overrides status color on active/on-hold lists",
     "Active list with age filter applied",
     "1. Go to /active; 2. Click '< 3 Days' filter",
     "Card border uses green (#43a047) from age color, not status color",
     "","Not Run","High",""),

    ("TC-RCD-07","RecordCard","Inconsistent Fields",
     "Pending records show Inconsistent Fields chips",
     "List has a pending record",
     "1. Find a pending card",
     "Divider and chip list of invalid field names visible below the primary info",
     "","Not Run","High",""),

    ("TC-RCD-08","RecordCard","Inconsistent Fields",
     "Accepted records do NOT show Inconsistent Fields section",
     "List has an accepted record",
     "1. Find an accepted card",
     "No Divider or chip list below the primary info row",
     "","Not Run","High",""),

    ("TC-RCD-09","RecordCard","Inconsistent Fields",
     "More than 8 inconsistent fields shows '+N more' text",
     "Record with >8 invalid fields",
     "1. Find or create a record with 9+ invalid fields; 2. Observe card",
     "First 8 chips shown; '+1 more' (or +N) text appears after",
     "","Not Run","Medium",""),

    ("TC-RCD-10","RecordCard","Interaction",
     "Clicking a card navigates to DetailsPage for that record",
     "List page loaded with records",
     "1. Click any record card",
     "Route changes to /<ExecutionId>; DetailsPage loads for that record",
     "","Not Run","Critical",""),

    ("TC-RCD-11","RecordCard","Interaction",
     "Hover on card shows scale(1.02) and enhanced shadow",
     "List page loaded",
     "1. Hover over a record card",
     "Card slightly enlarges (scale transform) and shadow intensifies",
     "","Not Run","Medium",""),

    # ── 9. DETAILS PAGE ──────────────────────────────────────────────────────
    ("TC-DET-01","DetailsPage","Loading",
     "Loader spinner shown while record details are fetching",
     "User clicks a record from list",
     "1. Click a record card from the list",
     "Full-page Loader component displayed until data arrives",
     "","Not Run","High",""),

    ("TC-DET-02","DetailsPage","Error Handling",
     "ErrorPage shown if details API call fails",
     "API is unreachable or returns non-2xx",
     "1. Simulate API failure; 2. Navigate to a record",
     "ErrorPage renders with error message and a Retry button",
     "","Not Run","High",""),

    ("TC-DET-03","DetailsPage","Error Handling",
     "Retry button on ErrorPage re-fetches data",
     "ErrorPage is displayed",
     "1. Click 'Retry' on ErrorPage",
     "fetchData() called again; Loader shows; data loads if API is now available",
     "","Not Run","High",""),

    ("TC-DET-04","DetailsPage","Snackbar",
     "Success snackbar shown after accepting a suggestion",
     "DetailsPage loaded with pending fields",
     "1. Select a suggestion; 2. Click Accept; 3. Confirm in dialog",
     "Green snackbar appears at top-center: 'Data accepted successfully'",
     "","Not Run","Critical",""),

    ("TC-DET-05","DetailsPage","Snackbar",
     "Error snackbar shown on Accept API failure",
     "DetailsPage loaded; Accept API returns error",
     "1. Select suggestion; 2. Attempt accept; 3. Simulate API error",
     "Red snackbar appears with error message from API",
     "","Not Run","High",""),

    ("TC-DET-06","DetailsPage","Snackbar",
     "Snackbar auto-dismisses after 3 seconds",
     "Snackbar is visible",
     "1. Trigger any snackbar; 2. Wait 3 seconds",
     "Snackbar disappears automatically",
     "","Not Run","Medium",""),

    ("TC-DET-07","DetailsPage","Snackbar",
     "Snackbar positioned at top-center",
     "Snackbar triggered",
     "1. Trigger success snackbar",
     "Snackbar appears at top-center of viewport",
     "","Not Run","Medium",""),

    # ── 10. EXECUTION INFO BOX ───────────────────────────────────────────────
    ("TC-EIB-01","ExecutionInfoBox","Layout",
     "Execution details displayed in a 3-column grid",
     "DetailsPage loaded with execution data",
     "1. Open a record's detail page",
     "ExecutionInfoBox grid shows 3 fields per row (md breakpoint and above)",
     "","Not Run","High",""),

    ("TC-EIB-02","ExecutionInfoBox","Layout",
     "Grid collapses to 1 column on small screens",
     "DetailsPage on a narrow viewport (<768px)",
     "1. Resize browser to mobile width; 2. View ExecutionInfoBox",
     "All fields stack vertically in a single column",
     "","Not Run","Medium",""),

    ("TC-EIB-03","ExecutionInfoBox","Display",
     "Empty field values show dash '—' instead of empty string",
     "Record with some null execution info fields",
     "1. Open any detail page; look for null-valued fields",
     "Null/undefined values display as '—'",
     "","Not Run","Medium",""),

    ("TC-EIB-04","ExecutionInfoBox","Display",
     "Section titled 'Execution Information' is shown",
     "DetailsPage loaded",
     "1. Observe DetailsPage top section",
     "h6 heading 'Execution Information' is visible above the data grid",
     "","Not Run","Low",""),

    # ── 11. CORRECTIONS TABLE ────────────────────────────────────────────────
    ("TC-COR-01","CorrectionsTable","Empty State",
     "Empty state shown when no invalid fields",
     "Record has no invalid fields (data=[]) ",
     "1. Open a fully accepted record with no invalid data",
     "Text 'No invalid fields found.' displayed centered",
     "","Not Run","High",""),

    ("TC-COR-02","CorrectionsTable","Group Accordion",
     "Groups are expanded by default",
     "DetailsPage loaded with corrections data",
     "1. Observe the corrections table immediately after load",
     "All field groups are expanded; suggestions are visible",
     "","Not Run","High",""),

    ("TC-COR-03","CorrectionsTable","Group Accordion",
     "Clicking group header collapses the group",
     "A group is expanded",
     "1. Click on the group header bar",
     "Suggestions and existing data section collapse with animation; icon changes to ExpandMore",
     "","Not Run","High",""),

    ("TC-COR-04","CorrectionsTable","Group Accordion",
     "Clicking collapsed group header expands it",
     "A group is collapsed",
     "1. Click the collapsed group header",
     "Group body expands with animation; ExpandLess icon shown",
     "","Not Run","High",""),

    ("TC-COR-05","CorrectionsTable","Status Chips",
     "'L0 Data' chip shown when field status is l0_data",
     "Record with L0 data field status",
     "1. Open a record where a field has L0 data rejection",
     "Red 'L0 Data' chip visible in group header",
     "","Not Run","Medium",""),

    ("TC-COR-06","CorrectionsTable","Status Chips",
     "'On Hold' chip shown when field status is on_hold",
     "Record with an on-hold field",
     "1. Open a record with an on-hold field",
     "Purple 'On Hold' chip visible in group header",
     "","Not Run","Medium",""),

    ("TC-COR-07","CorrectionsTable","Status Chips",
     "'Accepted' chip shown when field is already accepted",
     "Record with a previously accepted field",
     "1. Open a record with accepted correction",
     "Green 'Accepted' chip and 'Suggestion Selected'/'Custom Masterlist Dropdown Value Selected' chip visible",
     "","Not Run","High",""),

    ("TC-COR-08","CorrectionsTable","Existing Data Row",
     "Existing data bar shows current field values",
     "DetailsPage loaded with corrections",
     "1. Observe the gray bar at the top of each group",
     "'Existing Data' label and current field values shown; invalid field value appears in red (#ef4444)",
     "","Not Run","High",""),

    ("TC-COR-09","CorrectionsTable","Accept Button",
     "Accept button disabled when no suggestion is selected",
     "Group expanded, no suggestion selected",
     "1. Observe Accept button without selecting a suggestion",
     "Accept button is disabled (grayed out)",
     "","Not Run","Critical",""),

    ("TC-COR-10","CorrectionsTable","Accept Button",
     "Accept button enabled after selecting a suggestion",
     "Group expanded",
     "1. Click on a suggestion row (radio button)",
     "Accept button becomes enabled (green)",
     "","Not Run","Critical",""),

    ("TC-COR-11","CorrectionsTable","Accept Confirm Dialog",
     "Clicking Accept opens confirmation dialog",
     "A suggestion is selected",
     "1. Click the green 'Accept' button",
     "Modal dialog opens titled 'Confirm Accept' with field name and Yes/Cancel buttons",
     "","Not Run","Critical",""),

    ("TC-COR-12","CorrectionsTable","Accept Confirm Dialog",
     "'Yes, Accept' button shows loading spinner while accepting",
     "Confirm dialog open",
     "1. Click 'Yes, Accept'",
     "Button text changes to 'Accepting...' with circular spinner; button disabled during request",
     "","Not Run","High",""),

    ("TC-COR-13","CorrectionsTable","Accept Confirm Dialog",
     "Cancel closes dialog without submitting",
     "Confirm dialog open",
     "1. Click 'Cancel'",
     "Dialog closes; no API call made; selection preserved",
     "","Not Run","High",""),

    ("TC-COR-14","CorrectionsTable","Reject Button",
     "Reject All button disabled when a suggestion IS selected",
     "A suggestion is selected",
     "1. Select a suggestion; 2. Observe Reject All button",
     "Reject All button is disabled while a suggestion is selected",
     "","Not Run","High",""),

    ("TC-COR-15","CorrectionsTable","Reject Button",
     "Reject All enabled when no suggestion selected",
     "No suggestion selected",
     "1. Ensure no radio is selected; 2. Observe Reject All button",
     "Reject All button is enabled (red)",
     "","Not Run","High",""),

    ("TC-COR-16","CorrectionsTable","Reject Button",
     "Reject All hidden for non-pending fields",
     "Field status is accepted/L0/on-hold",
     "1. Open a record with accepted fields",
     "Accept and Reject All buttons are not rendered for accepted groups",
     "","Not Run","High",""),

    ("TC-COR-17","CorrectionsTable","No Suggestions",
     "'No suggestions available' shown when suggestions array is empty",
     "Group with empty suggestions list",
     "1. Open a record group with no suggestions",
     "Gray centered text 'No suggestions available' displayed in the suggestions area",
     "","Not Run","Medium",""),

    # ── 12. SUGGESTION ROW ───────────────────────────────────────────────────
    ("TC-SGR-01","SuggestionRow","Selection",
     "Clicking a suggestion row selects it (radio becomes checked)",
     "Group with suggestions, none selected",
     "1. Click any suggestion row",
     "Radio button in that row becomes checked; row background changes to selection color",
     "","Not Run","Critical",""),

    ("TC-SGR-02","SuggestionRow","Selection",
     "Clicking a selected suggestion deselects it",
     "A suggestion is selected",
     "1. Click the already-selected suggestion row",
     "Radio unchecks; row returns to default white background",
     "","Not Run","High",""),

    ("TC-SGR-03","SuggestionRow","Confidence Tooltip",
     "Tooltip shows confidence score on hover",
     "Group with suggestions that have a score field",
     "1. Hover over a suggestion row that has a score",
     "Tooltip shows 'Confidence: X%'",
     "","Not Run","Medium",""),

    ("TC-SGR-04","SuggestionRow","Accepted State",
     "Suggestions show dimmed/read-only state for accepted records",
     "Field already accepted",
     "1. Open an accepted record group",
     "All suggestion rows have opacity 0.5; cursor is 'not-allowed'; radio is disabled",
     "","Not Run","High",""),

    ("TC-SGR-05","SuggestionRow","EditableField",
     "coreCount field shows edit icon for VM suType when suggestion selected",
     "Record with sutType=vm; suggestion selected",
     "1. Select a suggestion in a vm record group with coreCount field",
     "Pencil (edit) icon appears next to the coreCount value",
     "","Not Run","High",""),

    ("TC-SGR-06","SuggestionRow","EditableField",
     "Clicking edit icon shows inline input for coreCount",
     "coreCount edit icon visible",
     "1. Click the pencil icon on coreCount field",
     "InputBase appears prefilled with current value; check and X icons appear",
     "","Not Run","High",""),

    ("TC-SGR-07","SuggestionRow","EditableField",
     "Pressing Enter saves the edited coreCount value",
     "Editing coreCount inline",
     "1. Change value; 2. Press Enter",
     "Field saves; view mode shows new value; input disappears",
     "","Not Run","High",""),

    ("TC-SGR-08","SuggestionRow","EditableField",
     "Pressing Escape cancels inline edit",
     "Editing coreCount inline",
     "1. Change value; 2. Press Escape",
     "Original value restored; input disappears",
     "","Not Run","High",""),

    ("TC-SGR-09","SuggestionRow","EditableField",
     "coreCount edit NOT available for non-vm sutType",
     "Record with sutType != 'vm'",
     "1. Select a suggestion; 2. Observe coreCount field",
     "No edit icon shown; field is display-only",
     "","Not Run","High",""),

    # ── 13. CHOOSE OTHER VALUE DROPDOWN ──────────────────────────────────────
    ("TC-CVD-01","ChooseOtherValueDropdown","Display",
     "Dropdown visible for pending field groups",
     "Field group in pending state",
     "1. Expand a pending group in CorrectionsTable",
     "Autocomplete dropdown labeled 'Choose other <fieldName>:' is visible",
     "","Not Run","High",""),

    ("TC-CVD-02","ChooseOtherValueDropdown","Display",
     "Dropdown hidden for non-pending (accepted/on-hold) field groups",
     "Field group in accepted state",
     "1. Expand an accepted group",
     "Choose Other Value dropdown is not rendered (returns null for !isPending)",
     "","Not Run","High",""),

    ("TC-CVD-03","ChooseOtherValueDropdown","Options Loading",
     "Opening dropdown fetches unique values from API",
     "Pending group dropdown visible",
     "1. Click/open the autocomplete dropdown",
     "Spinner appears; unique masterlist values loaded into dropdown options",
     "","Not Run","High",""),

    ("TC-CVD-04","ChooseOtherValueDropdown","Selection",
     "Selecting a value marks dropdown row as selected and fetches metadata",
     "Dropdown options are loaded",
     "1. Select a value from the dropdown",
     "Dropdown row border turns to accent color; 'Custom Value' label appears; metadata spinner shows briefly",
     "","Not Run","Critical",""),

    ("TC-CVD-05","ChooseOtherValueDropdown","Selection",
     "Clearing dropdown selection clears custom suggestion",
     "Custom value selected",
     "1. Clear the autocomplete value (X button)",
     "Custom suggestion row disappears; selection deactivated",
     "","Not Run","High",""),

    ("TC-CVD-06","ChooseOtherValueDropdown","Custom SuggestionRow",
     "Custom metadata renders as a SuggestionRow below dropdown",
     "Value selected from dropdown; metadata fetched",
     "1. Select a value; 2. Wait for metadata fetch",
     "A new SuggestionRow appears below the dropdown with the fetched metadata fields",
     "","Not Run","High",""),

    # ── 14. REJECT DIALOG ────────────────────────────────────────────────────
    ("TC-RJD-01","RejectDialog","Step 1 - Choose",
     "Clicking Reject All opens dialog with two option cards",
     "No suggestion selected; Reject All enabled",
     "1. Click 'Reject All' button",
     "Dialog opens titled 'Reject All Suggestions' with 'L0 Data' and 'Submit Draft Record' option cards",
     "","Not Run","Critical",""),

    ("TC-RJD-02","RejectDialog","Step 1 - Choose",
     "Option cards show hover lift animation",
     "Reject dialog at step 1",
     "1. Hover over 'L0 Data' card",
     "Card lifts (translateY -2px) with blue border and shadow",
     "","Not Run","Low",""),

    ("TC-RJD-03","RejectDialog","Step 2a - L0 Confirm",
     "L0 Data card navigates to confirmation step",
     "Reject dialog at step 1",
     "1. Click 'L0 Data' option card",
     "Dialog content changes to L0 confirmation screen with warning icon",
     "","Not Run","Critical",""),

    ("TC-RJD-04","RejectDialog","Step 2a - L0 Confirm",
     "Back button on L0 confirm returns to step 1",
     "Dialog on L0 confirm step",
     "1. Click back arrow button",
     "Returns to option choose step",
     "","Not Run","High",""),

    ("TC-RJD-05","RejectDialog","Step 2a - L0 Confirm",
     "Yes Confirm button submits L0 rejection",
     "Dialog on L0 confirm step",
     "1. Click 'Yes, Confirm'",
     "API called; dialog closes; page refreshes; snackbar 'Rejected due to L0 data'",
     "","Not Run","Critical",""),

    ("TC-RJD-06","RejectDialog","Step 2b - Draft Form",
     "Submit Draft Record card loads dynamic form fields",
     "Dialog at step 1",
     "1. Click 'Submit Draft Record'",
     "Dialog transitions to step 2 with loading spinner, then form fields rendered",
     "","Not Run","Critical",""),

    ("TC-RJD-07","RejectDialog","Step 2b - Draft Form",
     "Submit button disabled unless all fields are filled",
     "Draft form loaded with empty fields",
     "1. Observe Submit button without filling any field",
     "Submit button is disabled",
     "","Not Run","High",""),

    ("TC-RJD-08","RejectDialog","Step 2b - Draft Form",
     "Submit button enabled when all fields filled",
     "Draft form loaded",
     "1. Fill in all text fields",
     "Submit button becomes enabled",
     "","Not Run","High",""),

    ("TC-RJD-09","RejectDialog","Step 2b - Draft Form",
     "Submit button shows spinner during submission",
     "All form fields filled",
     "1. Click Submit",
     "Button shows CircularProgress spinner and 'Submitting...' text; button disabled",
     "","Not Run","High",""),

    ("TC-RJD-10","RejectDialog","Close",
     "X button closes dialog and resets to step 1",
     "Dialog open on any step",
     "1. Click the X (CloseIcon) button",
     "Dialog closes; reopening shows step 1 again",
     "","Not Run","High",""),

    # ── 15. ERROR PAGE ───────────────────────────────────────────────────────
    ("TC-ERR-01","ErrorPage","Display",
     "ErrorPage renders error message",
     "API returns error on list page",
     "1. Simulate API error on list page load",
     "ErrorPage renders with the specific message from the error",
     "","Not Run","High",""),

    ("TC-ERR-02","ErrorPage","Retry",
     "Retry button on ErrorPage triggers retry fetch",
     "ErrorPage visible",
     "1. Click 'Retry' button",
     "List re-fetches; if API is healthy, records load normally",
     "","Not Run","High",""),

    # ── 16. LOADER ───────────────────────────────────────────────────────────
    ("TC-LDR-01","Loader","Display",
     "Loader component renders centered spinner",
     "Any loading state",
     "1. Trigger any loading state (page load, accept, etc.)",
     "Centered spinner visible; no layout breakage",
     "","Not Run","Medium",""),

    # ── 17. RESPONSIVE DESIGN ────────────────────────────────────────────────
    ("TC-RES-01","Responsive","Mobile (<768px)",
     "Search bar width adapts to 90% on mobile",
     "List page on narrow viewport",
     "1. Resize to 375px width (iPhone SE); 2. Observe search input",
     "Search field width is ~90% of viewport; no horizontal overflow",
     "","Not Run","High",""),

    ("TC-RES-02","Responsive","Mobile (<768px)",
     "Filter buttons wrap on narrow screens",
     "Mobile viewport",
     "1. On mobile, check filter button row",
     "Filter buttons wrap to next line; no horizontal scroll (flexWrap: wrap)",
     "","Not Run","High",""),

    ("TC-RES-03","Responsive","Tablet (768–1200px)",
     "Record cards 65vw remain readable on tablets",
     "Tablet viewport",
     "1. Test list page at 768px wide",
     "Cards render at 65vw; text is readable; no clipping",
     "","Not Run","Medium",""),

    ("TC-RES-04","Responsive","Large Screen (>1200px)",
     "Layout uses available space appropriately on large screens",
     "1440px or wider viewport",
     "1. Open the app on a large monitor",
     "Cards centered at 65vw; comfortable padding on sides",
     "","Not Run","Low",""),

    # ── 18. CROSS-BROWSER ────────────────────────────────────────────────────
    ("TC-CBR-01","Cross-Browser","Chrome",
     "App renders and functions correctly on latest Chrome",
     "Chrome latest installed",
     "1. Open app in Chrome; 2. Test all main flows",
     "No rendering issues or JS errors in Chrome",
     "","Not Run","Critical",""),

    ("TC-CBR-02","Cross-Browser","Firefox",
     "App renders and functions correctly on latest Firefox",
     "Firefox latest installed",
     "1. Open app in Firefox; 2. Test main flows",
     "No rendering or functionality issues",
     "","Not Run","High",""),

    ("TC-CBR-03","Cross-Browser","Edge",
     "App renders correctly on Microsoft Edge",
     "Edge installed",
     "1. Open app in Edge; 2. Test main flows",
     "No issues on Edge",
     "","Not Run","Medium",""),

    # ── 19. PERFORMANCE / UX ─────────────────────────────────────────────────
    ("TC-PER-01","Performance","Initial Load",
     "List page shows first 50 records in under 3 seconds",
     "Fresh page load with backend running",
     "1. Open /active; 2. Measure time to first records render",
     "Records visible within 3 seconds",
     "","Not Run","High",""),

    ("TC-PER-02","Performance","Search Response",
     "Search results update within 1 second after debounce",
     "List page loaded",
     "1. Type a search term; wait 500ms debounce; observe",
     "Results update within 1 second of debounce completing",
     "","Not Run","High",""),

    ("TC-PER-03","Performance","Animations",
     "Card hover and filter button animations feel smooth (60fps)",
     "Any list page",
     "1. Hover over cards and filter buttons rapidly",
     "Transitions feel fluid; no jank or flickering",
     "","Not Run","Medium",""),

    ("TC-PER-04","Performance","Memory",
     "Navigating between list pages does not accumulate stale requests",
     "Auth/abort logic",
     "1. Quickly switch between /active and /completed several times",
     "Only the latest fetch results are shown; no ghost data from previous pages; no console AbortError crashes",
     "","Not Run","High",""),

    # ── 20. ACCESSIBILITY ────────────────────────────────────────────────────
    ("TC-ACC-01","Accessibility","Keyboard",
     "Search input is keyboard-accessible (Tab to focus, type to search)",
     "Any list page",
     "1. Tab to search input; 2. Type query",
     "Focus indicator visible; search works via keyboard",
     "","Not Run","High",""),

    ("TC-ACC-02","Accessibility","Keyboard",
     "Filter buttons activatable via keyboard",
     "List page",
     "1. Tab to a filter button; press Enter or Space",
     "Filter activates; records update",
     "","Not Run","High",""),

    ("TC-ACC-03","Accessibility","Contrast",
     "All text has sufficient color contrast (WCAG AA 4.5:1)",
     "Any page",
     "1. Run a contrast audit tool (e.g. Lighthouse) on the app",
     "Pass ratio ≥ AA for all informational text",
     "","Not Run","Medium",""),

    ("TC-ACC-04","Accessibility","ARIA",
     "Dialog has appropriate ARIA role and focus trap",
     "Any modal dialog open",
     "1. Open Accept or Reject dialog; 2. Check ARIA roles",
     "Dialog element has role='dialog'; focus trapped inside while open",
     "","Not Run","Medium",""),
]

# ─────────────────────────────────────────────────────────────────────────────
# BUILD WORKBOOK
# ─────────────────────────────────────────────────────────────────────────────

wb = openpyxl.Workbook()
wb.remove(wb.active)   # remove default sheet

# ── Sheet 1: Cover Page ─────────────────────────────────────────────────────
ws_cover = wb.create_sheet("Cover Page")
ws_cover.sheet_view.showGridLines = False
ws_cover.column_dimensions["A"].width = 5
ws_cover.column_dimensions["B"].width = 55
ws_cover.column_dimensions["C"].width = 35

def cw(r, c, v, bg=None, fg="FF000000", bold=False, sz=11, ha="left", wrap=False, italic=False):
    cell = ws_cover.cell(row=r, column=c, value=v)
    if bg: cell.fill = fill(bg)
    cell.font = font(size=sz, bold=bold, color=fg, italic=italic)
    cell.alignment = align(ha, wrap=wrap)
    return cell

# Title block
ws_cover.merge_cells("A1:C1"); ws_cover.row_dimensions[1].height = 8
ws_cover.merge_cells("A2:C2")
cw(2, 1, "AMD Data Hygiene — QA Test Report", bg=C["navy"], fg=C["white"], bold=True, sz=22, ha="center")
ws_cover.row_dimensions[2].height = 50

ws_cover.merge_cells("A3:C3")
cw(3, 1, "Frontend & UI Test Cases", bg=C["navy_light"], fg=C["white"], sz=13, ha="center", italic=True)
ws_cover.row_dimensions[3].height = 28

ws_cover.merge_cells("A4:C4"); ws_cover.row_dimensions[4].height = 12

# Info grid
info = [
    ("Project",      "AMD Data Hygiene"),
    ("Version",      "1.0"),
    ("Report Date",  str(date.today().strftime("%B %d, %Y"))),
    ("Prepared By",  "QA Team"),
    ("Branch",       "quality-analysis"),
    ("Framework",    "React + Vite + Material UI"),
    ("Test Scope",   "UI / Frontend — All Components & Pages"),
    ("Total TCs",    str(len(TESTS))),
]
for i, (k, v) in enumerate(info):
    r = 5 + i
    ws_cover.row_dimensions[r].height = 24
    cw(r, 1, "", bg=C["off_white"])
    cw(r, 2, k, bg=C["gray_bg"], bold=True, fg=C["navy"], sz=11)
    cw(r, 3, v, bg=C["off_white"], fg="FF1E293B", sz=11)

ws_cover.merge_cells("A14:C14"); ws_cover.row_dimensions[14].height = 16
ws_cover.merge_cells("A15:C15")
cw(15, 1, "STATUS LEGEND", bg=C["navy"], fg=C["white"], bold=True, sz=12, ha="center")
ws_cover.row_dimensions[15].height = 26

legend = [("Not Run", C["notrun_bg"], C["notrun_fg"]),
          ("Pass",    C["pass_bg"],   C["pass_fg"]),
          ("Fail",    C["fail_bg"],   C["fail_fg"]),
          ("Blocked", C["blocked_bg"],C["blocked_fg"]),
          ("N/A",     C["na_bg"],     C["na_fg"])]
for j, (s, bg, fg_c) in enumerate(legend):
    r = 16 + j
    ws_cover.row_dimensions[r].height = 22
    ws_cover.merge_cells(f"B{r}:C{r}")
    cw(r, 2, s, bg=bg, fg=fg_c, bold=True, sz=11, ha="center")

ws_cover.merge_cells("A22:C22"); ws_cover.row_dimensions[22].height = 16
ws_cover.merge_cells("A23:C23")
cw(23, 1, "PRIORITY LEGEND", bg=C["navy"], fg=C["white"], bold=True, sz=12, ha="center")
ws_cover.row_dimensions[23].height = 26
pri_legend = [("Critical","FF991B1B"),("High","FF9A3412"),("Medium","FF854D0E"),("Low","FF166534")]
for j, (p, c_) in enumerate(pri_legend):
    r = 24 + j
    ws_cover.row_dimensions[r].height = 22
    ws_cover.merge_cells(f"B{r}:C{r}")
    cc = ws_cover.cell(row=r, column=2, value=p)
    cc.fill  = fill(C["off_white"])
    cc.font  = font(bold=True, color=c_, size=11)
    cc.alignment = align("center")


# ── Sheet 2: Test Summary ────────────────────────────────────────────────────
ws_sum = wb.create_sheet("Test Summary")
ws_sum.sheet_view.showGridLines = False

MODULES = [
    "Navigation","List Pages","ListHeader","RecordList","RecordCard",
    "DetailsPage","ExecutionInfoBox","CorrectionsTable","SuggestionRow",
    "ChooseOtherValueDropdown","RejectDialog","ErrorPage","Loader",
    "Responsive","Cross-Browser","Performance","Accessibility"
]
status_opts = ["Not Run","Pass","Fail","Blocked","N/A"]

# Count per module per status
from collections import defaultdict
counts = defaultdict(lambda: defaultdict(int))
total_counts = defaultdict(int)
for tc in TESTS:
    mod = tc[1]
    st  = tc[8]
    counts[mod][st] += 1
    total_counts[st] += 1

# Title
ws_sum.merge_cells("A1:H1")
c1 = ws_sum.cell(row=1, column=1, value="Test Summary Dashboard")
c1.fill = fill(C["navy"]); c1.font = font(size=16, bold=True, color=C["white"])
c1.alignment = align("center"); ws_sum.row_dimensions[1].height = 40

# Headers
headers = ["Module","Not Run","Pass","Fail","Blocked","N/A","Total","% Pass"]
col_widths = [32, 12, 10, 10, 12, 10, 10, 12]
for ci, (h, w) in enumerate(zip(headers, col_widths), 1):
    ws_sum.column_dimensions[get_column_letter(ci)].width = w
    col_header(ws_sum, 2, ci, h)

# Data rows
for ri, mod in enumerate(MODULES, 3):
    nr  = counts[mod].get("Not Run", 0)
    pas = counts[mod].get("Pass",    0)
    fai = counts[mod].get("Fail",    0)
    blo = counts[mod].get("Blocked", 0)
    na  = counts[mod].get("N/A",     0)
    tot = nr + pas + fai + blo + na
    pct = f"{round(pas/tot*100)}%" if tot else "—"

    row_bg = C["off_white"] if ri % 2 == 0 else C["white"]
    write_cell(ws_sum, ri, 1, mod,    bg=row_bg, bold=True, fg=C["navy"])
    write_cell(ws_sum, ri, 2, nr,     bg=C["notrun_bg"],  fg=C["notrun_fg"], h_align="center")
    write_cell(ws_sum, ri, 3, pas,    bg=C["pass_bg"],    fg=C["pass_fg"],   h_align="center")
    write_cell(ws_sum, ri, 4, fai,    bg=C["fail_bg"],    fg=C["fail_fg"],   h_align="center")
    write_cell(ws_sum, ri, 5, blo,    bg=C["blocked_bg"], fg=C["blocked_fg"],h_align="center")
    write_cell(ws_sum, ri, 6, na,     bg=C["na_bg"],      fg=C["na_fg"],     h_align="center")
    write_cell(ws_sum, ri, 7, tot,    bg=row_bg, bold=True, fg=C["navy"],    h_align="center")
    write_cell(ws_sum, ri, 8, pct,    bg=row_bg, h_align="center",           fg=C["gray_text"])
    ws_sum.row_dimensions[ri].height = 22

# Totals row
tr = len(MODULES) + 3
write_cell(ws_sum, tr, 1, "TOTAL", bg=C["navy"], fg=C["white"], bold=True, size=11)
for ci, key in enumerate(["Not Run","Pass","Fail","Blocked","N/A"], 2):
    bg, fg_c = STATUS_STYLE[key]
    write_cell(ws_sum, tr, ci, total_counts[key], bg=bg, fg=fg_c, bold=True, h_align="center")
write_cell(ws_sum, tr, 7, len(TESTS), bg=C["navy"], fg=C["white"], bold=True, h_align="center")
ws_sum.row_dimensions[tr].height = 26

# ── Sheet 3: All Test Cases ───────────────────────────────────────────────────
ws_tc = wb.create_sheet("Test Cases")
ws_tc.sheet_view.showGridLines = False
ws_tc.freeze_panes = "A3"

TC_COLS = [
    ("TC ID",          10), ("Module",        22), ("Component",      26),
    ("Test Case",      50), ("Pre-conditions",32), ("Steps",          52),
    ("Expected Result",52), ("Actual Result", 40), ("Status",         13),
    ("Priority",       13), ("Remarks",       30),
]
col_letters = [get_column_letter(i+1) for i in range(len(TC_COLS))]
for i, (_, w) in enumerate(TC_COLS):
    ws_tc.column_dimensions[col_letters[i]].width = w

# Title
ws_tc.merge_cells(f"A1:{col_letters[-1]}1")
title_cell = ws_tc.cell(row=1, column=1, value="AMD Data Hygiene — UI/Frontend Test Cases")
title_cell.fill = fill(C["navy"])
title_cell.font = font(size=14, bold=True, color=C["white"])
title_cell.alignment = align("center")
ws_tc.row_dimensions[1].height = 36

# Column headers
for i, (h, _) in enumerate(TC_COLS):
    col_header(ws_tc, 2, i+1, h)
ws_tc.row_dimensions[2].height = 24

# Group tracking for section headers
last_module = None
data_row = 3

for tc in TESTS:
    tc_id, module, component, desc, pre, steps, expected, actual, status, priority, remarks = tc

    if module != last_module:
        # Section separator row
        ws_tc.merge_cells(start_row=data_row, start_column=1, end_row=data_row, end_column=len(TC_COLS))
        sep_cell = ws_tc.cell(row=data_row, column=1, value=f"  ▌ {module.upper()}")
        sep_cell.fill = fill(C["navy_light"])
        sep_cell.font = font(size=11, bold=True, color=C["white"])
        sep_cell.alignment = align("left")
        ws_tc.row_dimensions[data_row].height = 20
        data_row += 1
        last_module = module

    row_bg = C["off_white"] if data_row % 2 == 0 else C["white"]

    write_cell(ws_tc, data_row, 1,  tc_id,     bg=C["blue_light"], bold=True, fg=C["navy"], size=10)
    write_cell(ws_tc, data_row, 2,  module,    bg=row_bg, fg=C["navy"], bold=True, size=10)
    write_cell(ws_tc, data_row, 3,  component, bg=row_bg, size=10)
    write_cell(ws_tc, data_row, 4,  desc,      bg=row_bg, size=10, wrap=True)
    write_cell(ws_tc, data_row, 5,  pre,       bg=row_bg, size=10, wrap=True, fg=C["gray_text"], italic=True)
    write_cell(ws_tc, data_row, 6,  steps,     bg=row_bg, size=10, wrap=True)
    write_cell(ws_tc, data_row, 7,  expected,  bg=row_bg, size=10, wrap=True, fg="FF065F46")
    write_cell(ws_tc, data_row, 8,  actual,    bg=row_bg, size=10, wrap=True, fg=C["gray_text"], italic=True)

    apply_status(ws_tc.cell(row=data_row, column=9, value=status), status)
    apply_priority(ws_tc.cell(row=data_row, column=10, value=priority), priority)
    write_cell(ws_tc, data_row, 11, remarks,   bg=row_bg, size=10, italic=True, fg=C["gray_text"])

    ws_tc.row_dimensions[data_row].height = 60
    data_row += 1

# ── Sheet 4: Module-Wise View (grouped) ──────────────────────────────────────
def make_module_sheet(wb, module_name, module_tests):
    safe = module_name.replace("/","-").replace(" ","_")[:28]
    ws = wb.create_sheet(safe)
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A3"

    cols = [("TC ID",10),("Component",28),("Test Case",48),
            ("Steps",48),("Expected Result",48),("Status",12),("Priority",12),("Remarks",28)]
    for i,(h,w) in enumerate(cols):
        ws.column_dimensions[get_column_letter(i+1)].width = w

    ws.merge_cells(f"A1:{get_column_letter(len(cols))}1")
    t = ws.cell(row=1, column=1, value=f"{module_name} — Test Cases")
    t.fill = fill(C["navy"]); t.font = font(size=13,bold=True,color=C["white"])
    t.alignment = align("center"); ws.row_dimensions[1].height = 34

    for i,(h,_) in enumerate(cols): col_header(ws, 2, i+1, h)
    ws.row_dimensions[2].height = 22

    for dr, tc in enumerate(module_tests, 3):
        tc_id,_,comp,desc,_,steps,exp,_,status,pri,rem = tc
        bg = C["off_white"] if dr % 2==0 else C["white"]
        write_cell(ws,dr,1,tc_id,   bg=C["blue_light"],bold=True,fg=C["navy"],size=10)
        write_cell(ws,dr,2,comp,    bg=bg,size=10)
        write_cell(ws,dr,3,desc,    bg=bg,size=10,wrap=True)
        write_cell(ws,dr,4,steps,   bg=bg,size=10,wrap=True)
        write_cell(ws,dr,5,exp,     bg=bg,size=10,wrap=True,fg="FF065F46")
        apply_status(ws.cell(row=dr,column=6,value=status), status)
        apply_priority(ws.cell(row=dr,column=7,value=pri), pri)
        write_cell(ws,dr,8,rem,     bg=bg,size=10,italic=True,fg=C["gray_text"])
        ws.row_dimensions[dr].height = 55

# group by module
from collections import OrderedDict
by_module = OrderedDict()
for tc in TESTS:
    by_module.setdefault(tc[1], []).append(tc)

for mod, tcs in by_module.items():
    make_module_sheet(wb, mod, tcs)

# ── Sheet: Bug Tracker ────────────────────────────────────────────────────────
ws_bug = wb.create_sheet("Bug Tracker")
ws_bug.sheet_view.showGridLines = False
ws_bug.freeze_panes = "A3"

bug_cols = [
    ("Bug ID",12),("TC ID",12),("Module",22),("Component",26),("Bug Title",50),
    ("Description",50),("Severity",14),("Priority",12),("Status",16),
    ("Reported By",20),("Reported Date",18),("Fixed By",20),("Fixed Date",18),("Remarks",30),
]
for i,(h,w) in enumerate(bug_cols):
    ws_bug.column_dimensions[get_column_letter(i+1)].width = w

ws_bug.merge_cells(f"A1:{get_column_letter(len(bug_cols))}1")
bt = ws_bug.cell(row=1,column=1,value="Bug Tracker")
bt.fill=fill(C["navy"]); bt.font=font(size=14,bold=True,color=C["white"])
bt.alignment=align("center"); ws_bug.row_dimensions[1].height=36

for i,(h,_) in enumerate(bug_cols): col_header(ws_bug,2,i+1,h)
ws_bug.row_dimensions[2].height=22

# Severity dropdown hint row
SEVERITY = ["Blocker","Critical","Major","Minor","Trivial"]
BUG_STATUS = ["Open","In Progress","Fixed","Verified","Closed","Won't Fix"]

for r in range(3, 23):
    bg = C["off_white"] if r%2==0 else C["white"]
    for c in range(1, len(bug_cols)+1):
        cell = ws_bug.cell(row=r, column=c)
        cell.fill = fill(bg)
        cell.alignment = align(wrap=True)
        cell.border = border_all()
    ws_bug.row_dimensions[r].height = 30

# severity hint label
write_cell(ws_bug,3,7,"[Blocker/Critical/Major/Minor/Trivial]",
           bg=C["off_white"],fg=C["gray_text"],size=9,italic=True,wrap=True)
write_cell(ws_bug,3,9,"[Open/In Progress/Fixed/Verified/Closed]",
           bg=C["off_white"],fg=C["gray_text"],size=9,italic=True,wrap=True)

# ── Sheet: Execution Log ──────────────────────────────────────────────────────
ws_log = wb.create_sheet("Execution Log")
ws_log.sheet_view.showGridLines = False

log_cols = [
    ("Run #",8),("Date",16),("Tester",20),("Environment",20),
    ("Build/Branch",22),("TCs Run",10),("Passed",10),("Failed",10),
    ("Blocked",10),("Not Run",10),("Pass %",12),("Notes",50),
]
for i,(h,w) in enumerate(log_cols):
    ws_log.column_dimensions[get_column_letter(i+1)].width = w

ws_log.merge_cells(f"A1:{get_column_letter(len(log_cols))}1")
lt = ws_log.cell(row=1,column=1,value="Test Execution Log")
lt.fill=fill(C["navy"]); lt.font=font(size=14,bold=True,color=C["white"])
lt.alignment=align("center"); ws_log.row_dimensions[1].height=36

for i,(h,_) in enumerate(log_cols): col_header(ws_log,2,i+1,h)
ws_log.row_dimensions[2].height=22

for r in range(3, 18):
    bg = C["off_white"] if r%2==0 else C["white"]
    for c in range(1,len(log_cols)+1):
        cell = ws_log.cell(row=r,column=c)
        cell.fill = fill(bg)
        cell.alignment = align(wrap=True)
        cell.border = border_all()
    ws_log.row_dimensions[r].height = 28

# Example first row
run1_data = ["1", str(date.today().strftime("%Y-%m-%d")), "",
             "Local Dev", "quality-analysis",
             len(TESTS), 0, 0, 0, len(TESTS), "0%", "Initial run — all TCs pending"]
for ci, val in enumerate(run1_data, 1):
    write_cell(ws_log, 3, ci, val, bg=C["off_white"], size=10)


# ─────────────────────────────────────────────────────────────────────────────
# REORDER SHEETS
# ─────────────────────────────────────────────────────────────────────────────
sheet_order = ["Cover Page","Test Summary","Test Cases","Bug Tracker","Execution Log"]
module_sheets = [s for s in wb.sheetnames if s not in sheet_order]
final_order = sheet_order + module_sheets
for i, name in enumerate(final_order):
    wb.move_sheet(name, offset=i - wb.sheetnames.index(name))

# ─────────────────────────────────────────────────────────────────────────────
# SAVE
# ─────────────────────────────────────────────────────────────────────────────
OUT = r"d:\DH\Data-Hygiene\QA_Report_DataHygiene_UI.xlsx"
wb.save(OUT)
print(f"[OK] QA report saved -> {OUT}")
print(f"     Total test cases: {len(TESTS)}")
print(f"     Sheets: {wb.sheetnames}")

"""
QA Report Generator v2 — AMD Data Hygiene Frontend (post new-master-list merge)
Reflects all UI/frontend changes from the updated codebase.
"""

import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import date
from collections import defaultdict, OrderedDict

# ── Palette ──────────────────────────────────────────────────────────────────
C = {
    "navy":        "FF17233A",
    "navy_light":  "FF1E3A5F",
    "white":       "FFFFFFFF",
    "off_white":   "FFF8FAFD",
    "gray_bg":     "FFF1F5F9",
    "blue_light":  "FFD6E8FF",
    "header_row":  "FFE2ECF9",
    "gray_text":   "FF64748B",
    "pass_bg":     "FFD1FAE5", "pass_fg":    "FF065F46",
    "fail_bg":     "FFFEE2E2", "fail_fg":    "FF991B1B",
    "blocked_bg":  "FFFEF3C7", "blocked_fg": "FF92400E",
    "na_bg":       "FFF1F5F9", "na_fg":      "FF334155",
    "notrun_bg":   "FFF5F0FF", "notrun_fg":  "FF5B21B6",
    "p_crit":      "FFEF4444", "p_high":     "FFf97316",
    "p_med":       "FFEab308", "p_low":      "FF22C55E",
}

def fill(h):   return PatternFill("solid", fgColor=h)
def font(name="Calibri", size=11, bold=False, color="FF000000", italic=False):
    return Font(name=name, size=size, bold=bold, color=color, italic=italic)
def align(h="left", v="center", wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)
def bdr():
    s = Side(style="thin", color="FFD1D5DB")
    return Border(left=s, right=s, top=s, bottom=s)
def bdr_outer():
    s = Side(style="medium", color="FF94A3B8")
    return Border(left=s, right=s, top=s, bottom=s)

STATUS_STYLE = {
    "Pass":    (C["pass_bg"],    C["pass_fg"]),
    "Fail":    (C["fail_bg"],    C["fail_fg"]),
    "Blocked": (C["blocked_bg"], C["blocked_fg"]),
    "N/A":     (C["na_bg"],      C["na_fg"]),
    "Not Run": (C["notrun_bg"],  C["notrun_fg"]),
}
PRIORITY_COLOR = {
    "Critical": C["p_crit"], "High": C["p_high"],
    "Medium":   C["p_med"],  "Low":  C["p_low"],
}

def apply_status(cell, status):
    if status in STATUS_STYLE:
        bg, fg = STATUS_STYLE[status]
        cell.fill = fill(bg)
        cell.font = font(bold=True, color=fg, size=10)
    cell.alignment = align("center"); cell.border = bdr()

def apply_priority(cell, pri):
    c_ = PRIORITY_COLOR.get(pri, C["na_fg"])
    cell.fill = fill(C["white"]); cell.font = font(bold=True, color=c_, size=10)
    cell.alignment = align("center"); cell.border = bdr()

def wc(ws, row, col, value, bg=None, fg="FF000000", bold=False, size=11,
        h_align="left", wrap=False, italic=False):
    cell = ws.cell(row=row, column=col, value=value)
    if bg: cell.fill = fill(bg)
    cell.font  = font(size=size, bold=bold, color=fg, italic=italic)
    cell.alignment = align(h_align, wrap=wrap); cell.border = bdr()
    return cell

def ch(ws, row, col, value):
    cell = ws.cell(row=row, column=col, value=value)
    cell.fill = fill(C["header_row"]); cell.font = font(size=10, bold=True, color=C["navy"])
    cell.alignment = align("center"); cell.border = bdr(); return cell

# ─────────────────────────────────────────────────────────────────────────────
# TEST CASES
# Columns: TC_ID | Module | Component | Description | Pre-conditions |
#          Steps | Expected Result | Actual Result | Status | Priority | Remarks
# ─────────────────────────────────────────────────────────────────────────────
TESTS = [

    # ═══════════════════════════════════════════════════════
    # 1. NAVIGATION & ROUTING
    # ═══════════════════════════════════════════════════════
    ("TC-NAV-01","Navigation","Navbar",
     "Logo 'AMD_DH' click navigates to '/' (landing page)",
     "App loaded on any route",
     "1. Click the 'AMD_DH' logo in the top-left of the navbar",
     "Route changes to '/'; page title shows 'Data Hygiene Dashboard'",
     "","Not Run","High",""),

    ("TC-NAV-02","Navigation","Navbar",
     "'Active List' nav link navigates to /active",
     "App loaded",
     "1. Click 'Active List' in the navbar",
     "Route changes to '/active'; title shows 'My Active List'",
     "","Not Run","High",""),

    ("TC-NAV-03","Navigation","Navbar",
     "'Completed List' nav link navigates to /completed",
     "App loaded",
     "1. Click 'Completed List'",
     "Route changes to '/completed'; title shows 'My Completed List'",
     "","Not Run","High",""),

    ("TC-NAV-04","Navigation","Navbar",
     "'On Hold' nav link navigates to /on-hold",
     "App loaded",
     "1. Click 'On Hold'",
     "Route changes to '/on-hold'; title shows 'On Hold Records'",
     "","Not Run","High",""),

    ("TC-NAV-05","Navigation","Navbar",
     "Active route NavLink shows underline indicator",
     "User is on /active",
     "1. Observe navbar links while on /active",
     "'Active List' has bottomBorder underline; others are dimmed (opacity 0.7)",
     "","Not Run","Medium",""),

    ("TC-NAV-06","Navigation","Navbar",
     "Navbar stays fixed at top on scroll",
     "Any list page with many records",
     "1. Scroll down past the viewport height; 2. Observe navbar position",
     "Navbar remains pinned to top (position: fixed; AppBar)",
     "","Not Run","Medium",""),

    ("TC-NAV-07","Navigation","Routing",
     "Routes use stable 'key' props — navigating between modes resets state",
     "App loaded on any list page",
     "1. Go to /active; type in search; 2. Click 'Completed List'",
     "State resets — search clears, filter resets (key prop forces remount per route)",
     "","Not Run","High","Old routes did not have key props; now they do"),

    ("TC-NAV-08","Navigation","Routing",
     "/:id route opens DetailsPage for that record",
     "A valid ExecutionId is known",
     "1. Navigate directly to /<valid-id>",
     "DetailsPage loads with Execution Information and Corrections Table for that record",
     "","Not Run","High",""),

    ("TC-NAV-09","Navigation","Routing",
     "/completed/:id route removed — all details now use /:id",
     "App loaded",
     "1. Navigate to /completed/<id>",
     "Route does NOT match — should fall back (no separate completed/:id route exists anymore)",
     "","Not Run","Medium","Route removed in updated App.jsx"),

    # ═══════════════════════════════════════════════════════
    # 2. LIST PAGES
    # ═══════════════════════════════════════════════════════
    ("TC-LST-01","List Pages","Landing ('/')",
     "Landing page renders 'Data Hygiene Dashboard' title",
     "App at '/'",
     "1. Open '/'",
     "H3 heading 'Data Hygiene Dashboard' visible below the navbar",
     "","Not Run","High",""),

    ("TC-LST-02","List Pages","Landing ('/')",
     "Landing shows status filter buttons: Pending, Accepted, L0 Data, On Hold",
     "App at '/'",
     "1. Open '/'",
     "Four filter buttons visible: Pending (orange), Accepted (green), L0 Data (red), On Hold (yellow)",
     "","Not Run","High","'Rejected' is now labelled 'L0 Data' in the filter UI"),

    ("TC-LST-03","List Pages","Landing ('/')",
     "Clicking 'L0 Data' filter fetches records with status=rejected",
     "Landing page loaded",
     "1. Click 'L0 Data' filter button; 2. Observe API request",
     "API called with status=rejected; list updates to show L0 Data records",
     "","Not Run","Critical","Filter label changed from 'Rejected' to 'L0 Data'; API value still 'rejected'"),

    ("TC-LST-04","List Pages","Landing ('/')",
     "Filter toggle: clicking active filter deactivates it",
     "A filter button is active",
     "1. Click the active filter button again",
     "Filter deactivates; all records shown again",
     "","Not Run","High",""),

    ("TC-LST-05","List Pages","Active ('/active')",
     "Active list shows age filter buttons only",
     "User on /active",
     "1. Open /active",
     "Three age filter buttons visible: '< 3 Days' (green), '3 - 6 Days' (amber), '> 6 Days' (red); no status filters",
     "","Not Run","High",""),

    ("TC-LST-06","List Pages","Active ('/active')",
     "Age filter '<3 Days' sends age=green to API",
     "/active loaded",
     "1. Click '< 3 Days'; 2. Check network request",
     "API called with status=pending&age=green",
     "","Not Run","High",""),

    ("TC-LST-07","List Pages","Completed ('/completed')",
     "Completed list shows only Accepted and L0 Data filter buttons",
     "User on /completed",
     "1. Open /completed",
     "Only 'Accepted' and 'L0 Data' filter buttons shown (Pending and On Hold absent)",
     "","Not Run","High","allowedFilters restricts to [accepted, rejected]"),

    ("TC-LST-08","List Pages","Completed ('/completed')",
     "Completed list default fetches status=accepted,rejected",
     "User on /completed; no filter selected",
     "1. Open /completed; 2. Inspect API request",
     "API called with status=accepted,rejected",
     "","Not Run","Critical",""),

    ("TC-LST-09","List Pages","On Hold ('/on-hold')",
     "On Hold list shows age filter buttons",
     "User on /on-hold",
     "1. Open /on-hold",
     "Age filter buttons visible; default API call uses status=On Hold",
     "","Not Run","High",""),

    ("TC-LST-10","List Pages","Empty State",
     "Empty state message varies by filter active state",
     "Filter applied with 0 results",
     "1. Click a filter that returns no records",
     "Shows 'No records match the selected filter.'",
     "","Not Run","High",""),

    # ═══════════════════════════════════════════════════════
    # 3. LIST HEADER
    # ═══════════════════════════════════════════════════════
    ("TC-HDR-01","ListHeader","Search Bar",
     "Search placeholder text is correct",
     "Any list page loaded",
     "1. Click inside the search field without typing",
     "Placeholder text: 'Search Execution ID, Type, or Category...'",
     "","Not Run","Low",""),

    ("TC-HDR-02","ListHeader","Search Bar",
     "Typing triggers debounced API call after 500ms",
     "List page loaded",
     "1. Type 'AMD' quickly; 2. Wait 600ms",
     "Single API call fires ~500ms after typing stops; spinner appears during debounce",
     "","Not Run","High",""),

    ("TC-HDR-03","ListHeader","Search Bar",
     "Search field renders with gray border (#747d8a) and rounded corners (8px)",
     "Any list page",
     "1. Inspect search field styling",
     "Field has border: '1px solid #747d8aff'; borderRadius: 8px",
     "","Not Run","Medium","Border styling updated from old version"),

    ("TC-HDR-04","ListHeader","Search Bar",
     "CircularProgress spinner appears in search field end adornment while loading",
     "List page; search triggered",
     "1. Start typing; observe before debounce fires",
     "Small circular spinner (size=18) visible at right end of search field",
     "","Not Run","Medium",""),

    ("TC-HDR-05","ListHeader","Filter Buttons",
     "Filter button inactive state: white background, gray (#727579) border",
     "Any list page; no filter active",
     "1. Observe filter buttons at rest",
     "All filter buttons have white (#ffffff) background and gray (#727579ff) border",
     "","Not Run","Medium","Updated patchSx styling"),

    ("TC-HDR-06","ListHeader","Filter Buttons",
     "Filter button active state: colored background, white text, colored shadow",
     "A filter button is clicked",
     "1. Click any filter button",
     "Background fills with activeBg color; text turns white; shadow appears",
     "","Not Run","High",""),

    ("TC-HDR-07","ListHeader","Filter Buttons",
     "Filter button hover: slight lift (translateY -1px) and border color change",
     "List page loaded",
     "1. Hover over an inactive filter button",
     "Button lifts 1px and border color changes to the filter's accent color",
     "","Not Run","Medium","Updated hover effect: -1px (was -2px)"),

    ("TC-HDR-08","ListHeader","Filter Buttons — Updated Labels",
     "'Rejected' filter renamed to 'L0 Data' on Landing and Completed pages",
     "Landing or Completed page",
     "1. Observe status filter buttons",
     "Button shows label 'L0 Data' (not 'Rejected'); clicking it queries status=rejected",
     "","Not Run","Critical","Key change in v2"),

    ("TC-HDR-09","ListHeader","Filter Buttons — Updated Labels",
     "'On Hold' button on Landing page now uses yellow (#ca8a04 / #dbbc23) color scheme",
     "Landing page",
     "1. Observe the 'On Hold' filter button",
     "On Hold button accent is yellow (#dbbc23 active, #ca8a04 border); was purple in v1",
     "","Not Run","High","Color scheme changed from purple to yellow"),

    ("TC-HDR-10","ListHeader","Filter Buttons — Age",
     "Age filter colors updated: <3 Days uses emerald green (#10b981 active)",
     "/active or /on-hold page",
     "1. Click '< 3 Days' filter",
     "Active state background is #10b981 (emerald green), dot turns white",
     "","Not Run","Medium","AGE_FILTERS colors refreshed in v2"),

    # ═══════════════════════════════════════════════════════
    # 4. RECORD LIST (Virtualization)
    # ═══════════════════════════════════════════════════════
    ("TC-VRL-01","RecordList","Initial Load",
     "Full-page Loader shown when list is empty and loading",
     "Fresh page navigation",
     "1. Open any list page",
     "Loader spinner shown immediately; disappears once first records arrive",
     "","Not Run","High",""),

    ("TC-VRL-02","RecordList","Infinite Scroll",
     "Scrolling near bottom loads next page of records",
     "First page (50 records) loaded; more pages exist",
     "1. Scroll to near the bottom",
     "Additional records appended; bottom loader briefly visible",
     "","Not Run","Critical",""),

    ("TC-VRL-03","RecordList","Count Label",
     "'Total records: N' label shown when showCount is true",
     "List page loaded",
     "1. Observe above the record list",
     "'Total records: <N>' displayed for pages with age or status filters",
     "","Not Run","Medium",""),

    ("TC-VRL-04","RecordList","Row Heights",
     "Accepted/Approved cards: height 100; pending cards: height 160",
     "Mixed-status list",
     "1. Compare card heights in a list with both accepted and pending records",
     "Accepted cards visually shorter (no Inconsistent Fields section); pending cards taller",
     "","Not Run","Medium",""),

    ("TC-VRL-05","RecordList","Empty State",
     "Empty message shown when no records and not loading",
     "Filter with 0 results",
     "1. Apply filter returning 0 records; 2. Wait for load to finish",
     "Centered h6 message appears: 'No records match the selected filter.'",
     "","Not Run","High",""),

    # ═══════════════════════════════════════════════════════
    # 5. RECORD CARD
    # ═══════════════════════════════════════════════════════
    ("TC-RCD-01","RecordCard","Display",
     "Card displays ExecutionID, BenchmarkCategory, BenchmarkType, Status",
     "List with records",
     "1. Observe any record card",
     "All four data labels and values visible",
     "","Not Run","Critical",""),

    ("TC-RCD-02","RecordCard","Status Label — L0 Data",
     "Records with status='rejected' show 'L0 Data' as the status text on the card",
     "List has a rejected record",
     "1. Find a record with Status='Rejected'",
     "Card status field shows 'L0 Data' (not 'Rejected') — display renamed in RecordCard",
     "","Not Run","Critical","Key UX change in v2: record.Status === 'rejected' -> displays 'L0 Data'"),

    ("TC-RCD-03","RecordCard","Card Background",
     "All cards have white (#ffffff) background regardless of status",
     "List with mixed-status records",
     "1. Compare card backgrounds across statuses",
     "All cards are #ffffff; differentiation comes from left border color only",
     "","Not Run","High","Background no longer uses status tinted colors; simplified to white"),

    ("TC-RCD-04","RecordCard","Card Border — Pending",
     "Pending card has orange (#ffae00) left border (6px) and outer border (1.3px)",
     "List with pending records",
     "1. Observe pending card borders",
     "6px left border (#ffae00); 1.3px full border (#ffae00); white background",
     "","Not Run","High",""),

    ("TC-RCD-05","RecordCard","Card Border — Accepted",
     "Accepted card has green (#10b981) border",
     "List with accepted records",
     "1. Observe accepted card borders",
     "Green (#10b981) 6px left border; 1.3px outer border",
     "","Not Run","High",""),

    ("TC-RCD-06","RecordCard","Card Border — On Hold",
     "On Hold card has yellow (#f1e60b) border",
     "List with on-hold records",
     "1. Observe on-hold card borders",
     "Yellow (#f1e60b) 6px left border; 1.3px full border",
     "","Not Run","High","Changed to yellow from purple in v2"),

    ("TC-RCD-07","RecordCard","Card Border — L0 Data (Rejected)",
     "L0 Data (rejected) card has red (#ef4444) border",
     "List with rejected records",
     "1. Find a record with Status=Rejected (shown as L0 Data)",
     "Red (#ef4444) 6px left border",
     "","Not Run","High",""),

    ("TC-RCD-08","RecordCard","Card Hover",
     "Card hover: translateY(-4px) lift with shadow",
     "Any list page",
     "1. Hover over a card",
     "Card lifts 4px (was 4px scale in v1) and shows colored shadow; cursor changes to pointer",
     "","Not Run","Medium","Hover changed from scale(1.02) to translateY(-4px)"),

    ("TC-RCD-09","RecordCard","Inconsistent Fields",
     "Non-completed records show Inconsistent Fields chip list",
     "List has pending or on-hold records",
     "1. Observe a pending card",
     "Divider below main info; red error chips listing invalid field names",
     "","Not Run","High",""),

    ("TC-RCD-10","RecordCard","Suggestions Available Badge",
     "Pending records with suggestions show 'Suggestions available' italic label",
     "List has a pending record with suggestionsCount=true",
     "1. Find a pending card that has suggestions",
     "Italic text 'Suggestions available' visible in the Inconsistent Fields section (space-between row with chips)",
     "","Not Run","High","New in v2: SuggestionsCount prop passed to InconsistentFieldsList"),

    ("TC-RCD-11","RecordCard","Suggestions Available Badge",
     "Pending records without suggestions do NOT show 'Suggestions available'",
     "Pending record with suggestionsCount=false",
     "1. Find a pending card with no suggestions",
     "No 'Suggestions available' label present",
     "","Not Run","High",""),

    ("TC-RCD-12","RecordCard","Accepted/Approved",
     "Accepted/Approved cards do NOT show Inconsistent Fields section",
     "List with accepted records",
     "1. Find accepted card",
     "No divider or chip list below main info row",
     "","Not Run","High",""),

    ("TC-RCD-13","RecordCard","Navigation",
     "Clicking a card navigates to DetailsPage for that record",
     "List with records",
     "1. Click any card",
     "Route changes to /<ExecutionId>; DetailsPage loads",
     "","Not Run","Critical",""),

    # ═══════════════════════════════════════════════════════
    # 6. DETAILS PAGE
    # ═══════════════════════════════════════════════════════
    ("TC-DET-01","DetailsPage","Loading",
     "Full Loader shown while fetching record details",
     "User clicks a record card",
     "1. Click a card; observe immediately",
     "Full-page Loader spinner shown until data arrives",
     "","Not Run","High",""),

    ("TC-DET-02","DetailsPage","Error Handling",
     "ErrorPage with Retry shown on API failure",
     "API unreachable",
     "1. Simulate API failure; open a record",
     "ErrorPage component renders with error message and Retry button",
     "","Not Run","High",""),

    ("TC-DET-03","DetailsPage","Fallback",
     "Plain text fallback if executionData is null",
     "API returns no execution_details",
     "1. Response has no execution_details field",
     "Text 'No execution data to show' rendered (not a crash)",
     "","Not Run","Medium",""),

    ("TC-DET-04","DetailsPage","Snackbar — Success",
     "Green success snackbar appears at top-center after accepting a suggestion",
     "DetailsPage with pending corrections",
     "1. Select suggestion; 2. Click Accept; 3. Confirm in dialog",
     "Snackbar: 'Data accepted successfully' — green filled, top-center",
     "","Not Run","Critical",""),

    ("TC-DET-05","DetailsPage","Snackbar — Error",
     "Red error snackbar appears on API failure during accept",
     "Accept API returns error",
     "1. Simulate API error; attempt accept",
     "Red snackbar with API error message",
     "","Not Run","High",""),

    ("TC-DET-06","DetailsPage","Snackbar — Draft",
     "Success snackbar shown after successful draft record submission",
     "DetailsPage; draft dialog open",
     "1. Fill draft form; submit",
     "Green snackbar: 'Draft record submitted successfully'",
     "","Not Run","High","New flow in v2"),

    ("TC-DET-07","DetailsPage","Snackbar — L0",
     "Success snackbar shown after sending to L0",
     "L0 confirm dialog triggered",
     "1. Click 'Send to L0'; confirm",
     "Green snackbar: 'Rejected due to L0 data'",
     "","Not Run","High",""),

    ("TC-DET-08","DetailsPage","Snackbar",
     "Snackbar auto-dismisses after 3 seconds",
     "Any snackbar visible",
     "1. Trigger any action snackbar; wait 3s",
     "Snackbar disappears automatically",
     "","Not Run","Medium",""),

    # ═══════════════════════════════════════════════════════
    # 7. EXECUTION INFO BOX
    # ═══════════════════════════════════════════════════════
    ("TC-EIB-01","ExecutionInfoBox","Layout",
     "Execution info displays in a 3-column grid on desktop",
     "DetailsPage loaded",
     "1. Open a record's detail page on desktop",
     "ExecutionInfoBox shows 3 fields per row",
     "","Not Run","High",""),

    ("TC-EIB-02","ExecutionInfoBox","Layout",
     "Grid collapses to single column on mobile (<md breakpoint)",
     "Mobile viewport",
     "1. Resize to 375px width",
     "Fields stack vertically",
     "","Not Run","Medium",""),

    ("TC-EIB-03","ExecutionInfoBox","Display",
     "Null/undefined field values display as '-'",
     "Record with some null fields",
     "1. Open record with null execution info fields",
     "Null values show '-' (hyphen)",
     "","Not Run","Medium",""),

    # ═══════════════════════════════════════════════════════
    # 8. CORRECTIONS TABLE (Major v2 Changes)
    # ═══════════════════════════════════════════════════════
    ("TC-COR-01","CorrectionsTable","Empty State",
     "Shows 'No invalid fields found.' when data is empty",
     "Record with no invalid fields",
     "1. Open a fully-clean record",
     "Centered gray text 'No invalid fields found.' shown",
     "","Not Run","High",""),

    ("TC-COR-02","CorrectionsTable","Group Expansion",
     "Pending groups auto-expand; non-pending groups auto-collapse on load",
     "DetailsPage loaded with mixed-status groups",
     "1. Open a record with both pending and accepted groups",
     "Pending groups are expanded; accepted/on-hold/L0 groups are collapsed by default",
     "","Not Run","Critical","New in v2: expandedGroups initialized based on isPending status"),

    ("TC-COR-03","CorrectionsTable","Group Accordion",
     "Clicking collapsed group header expands it",
     "A group is collapsed",
     "1. Click on the group header",
     "Group body expands with collapse animation; icon changes to ExpandLess",
     "","Not Run","High",""),

    ("TC-COR-04","CorrectionsTable","Group Accordion",
     "Clicking expanded group header collapses it",
     "A group is expanded",
     "1. Click on the group header",
     "Group body collapses; icon changes to ExpandMore",
     "","Not Run","High",""),

    ("TC-COR-05","CorrectionsTable","Status Chips — Pending",
     "'Pending' chip shown in group header for pending/invalid fields",
     "A pending group",
     "1. Open a pending record, look at group header",
     "Orange 'Pending' chip visible (#fff7ed background, #9a3412 text)",
     "","Not Run","High","New in v2: Pending chip added to group header"),

    ("TC-COR-06","CorrectionsTable","Status Chips — L0 Data",
     "'L0 Data' red chip in header when field status is 'l0 data'",
     "Record with L0 data field",
     "1. Open record with L0 field",
     "Red 'L0 Data' chip (#fef2f2 bg, #dc2626 text, #fca5a5 border) in header",
     "","Not Run","Medium",""),

    ("TC-COR-07","CorrectionsTable","Status Chips — On Hold",
     "'On Hold' chip is now yellow (not purple) in header",
     "Record with On Hold field",
     "1. Open on-hold record group header",
     "Yellow 'On Hold' chip (#fef9c3 bg, #ca8a04 text, #fde047 border)",
     "","Not Run","High","Color changed from purple to yellow in v2"),

    ("TC-COR-08","CorrectionsTable","Status Chips — Accepted",
     "'Accepted' chip + 'Suggestion Selected'/'Custom...' chip shown for accepted fields",
     "Record with accepted correction",
     "1. Open accepted group header",
     "Green 'Accepted' chip and secondary chip indicating selection type",
     "","Not Run","High",""),

    ("TC-COR-09","CorrectionsTable","Suggestions Header",
     "'Suggestions' label only shown when suggestions exist; includes italic hint for pending",
     "Pending group with suggestions",
     "1. Open pending group with suggestions",
     "Label reads: 'Suggestions (Hover over an option to see confidence score)'",
     "","Not Run","Medium","New hint text added in v2"),

    ("TC-COR-10","CorrectionsTable","Suggestions Header",
     "Hints text NOT shown for accepted/on-hold groups",
     "Accepted group with suggestions",
     "1. Open accepted group",
     "No italic hint text in the Suggestions label",
     "","Not Run","Medium",""),

    ("TC-COR-11","CorrectionsTable","Draft Records Section",
     "Draft Record section shown below suggestions when group.draft_records exists",
     "On Hold record with draft data",
     "1. Open an on-hold record that has a draft_records object",
     "'Draft Record' label and a read-only SuggestionRow (no radio) displayed with ON_HOLD_THEME",
     "","Not Run","Critical","New in v2: draft_records display for On Hold fields"),

    ("TC-COR-12","CorrectionsTable","Draft Records Section",
     "Draft Record SuggestionRow has no radio button (showRadio=false)",
     "On Hold record with draft data",
     "1. Inspect Draft Record row",
     "No radio button on the left; row is read-only display",
     "","Not Run","High","showRadio prop newly added in v2"),

    ("TC-COR-13","CorrectionsTable","Draft Records Section",
     "VM draft records show History CPU(s) value override",
     "VM record, On Hold, with CPU history change",
     "1. Open VM on-hold record with CPU(s) history",
     "Draft record SuggestionRow shows CPU(s) value from history, not from draft_records",
     "","Not Run","High","VM CPU(s) override logic applied to draft display"),

    ("TC-COR-14","CorrectionsTable","Action Buttons",
     "Three action buttons shown for pending groups: Accept, Submit Draft Record, Send to L0",
     "Pending group; no suggestion selected",
     "1. Open pending group, observe action bar",
     "Green 'Accept', yellow 'Submit Draft Record', red 'Send to L0' buttons visible",
     "","Not Run","Critical","Old 'Reject All' replaced by two separate actions in v2"),

    ("TC-COR-15","CorrectionsTable","Action Buttons",
     "'Accept' button disabled when no suggestion selected",
     "Pending group, nothing selected",
     "1. Observe Accept button",
     "Accept button is disabled (grayed out)",
     "","Not Run","Critical",""),

    ("TC-COR-16","CorrectionsTable","Action Buttons",
     "'Submit Draft Record' and 'Send to L0' disabled when a suggestion IS selected",
     "A suggestion is selected",
     "1. Select any suggestion; observe the other two buttons",
     "Submit Draft Record and Send to L0 are disabled",
     "","Not Run","Critical","Mirrors old 'Reject All' disabled logic"),

    ("TC-COR-17","CorrectionsTable","Action Buttons",
     "All action buttons hidden for non-pending fields",
     "Accepted/On Hold/L0 field group",
     "1. Open non-pending group",
     "No action button bar rendered",
     "","Not Run","High",""),

    ("TC-COR-18","CorrectionsTable","Accept Confirm Dialog",
     "Clicking Accept opens 'Confirm Accept' dialog",
     "Suggestion selected in pending group",
     "1. Click 'Accept'",
     "Dialog opens with green check icon, field name highlighted, Cancel and 'Yes, Accept' buttons",
     "","Not Run","Critical",""),

    ("TC-COR-19","CorrectionsTable","Accept Confirm Dialog",
     "'Yes, Accept' shows spinner + 'Accepting...' while API is in-flight",
     "Confirm dialog open",
     "1. Click 'Yes, Accept'",
     "Button shows CircularProgress spinner with 'Accepting...' text; button disabled",
     "","Not Run","High",""),

    ("TC-COR-20","CorrectionsTable","L0 Confirm Dialog",
     "Clicking 'Send to L0' opens standalone L0 Confirm dialog",
     "No suggestion selected; Send to L0 enabled",
     "1. Click 'Send to L0'",
     "Dialog: 'Confirm Send to L0', warning amber icon, 'Yes, Confirm' (warning color) and Cancel",
     "","Not Run","Critical","L0 dialog now standalone in CorrectionsTable (not nested in RejectDialog)"),

    ("TC-COR-21","CorrectionsTable","L0 Confirm Dialog",
     "L0 dialog: clicking X or Cancel during submission is disabled",
     "L0 submission in progress",
     "1. Click 'Yes, Confirm'; immediately try to close",
     "Dialog cannot be closed while submitting (disabled state)",
     "","Not Run","High",""),

    ("TC-COR-22","CorrectionsTable","Draft Dialog",
     "Clicking 'Submit Draft Record' opens Draft Record dialog with dynamic fields",
     "No suggestion selected; Submit Draft enabled",
     "1. Click 'Submit Draft Record'",
     "Dialog opens titled 'Submit Draft Record'; spinner shown while fields load from API; fields appear after",
     "","Not Run","Critical",""),

    ("TC-COR-23","CorrectionsTable","Draft Dialog",
     "Draft dialog pre-fills field values from history changes",
     "Record with history data for the field",
     "1. Open Draft dialog for a field that has history",
     "Form fields pre-populated with history values (from buildDisplayObjectFromChanges)",
     "","Not Run","High","New in v2: draftInitialValues from history"),

    ("TC-COR-24","CorrectionsTable","Draft Dialog",
     "Integer-type fields render as number inputs with step=1, min=0",
     "Draft dialog open with integer fields",
     "1. Open draft dialog; inspect integer fields",
     "Text field type='number', inputProps step=1 min=0; label shows '(integer)' hint",
     "","Not Run","High","New in v2: integer fields get type=number"),

    ("TC-COR-25","CorrectionsTable","Draft Dialog",
     "Submit disabled until all fields are non-empty",
     "Draft dialog with empty fields",
     "1. Observe Submit button",
     "Submit button disabled; enables only after all fields have non-whitespace values",
     "","Not Run","High",""),

    ("TC-COR-26","CorrectionsTable","Draft Dialog",
     "Integer validation: error snackbar if non-integer value entered in integer field",
     "Draft dialog; integer field present",
     "1. Type '2.5' in an integer field; 2. Click Submit",
     "Error snackbar: 'The following fields require whole-number (integer) values: <fieldname>'",
     "","Not Run","High","New in v2: client-side integer validation before submit"),

    ("TC-COR-27","CorrectionsTable","Draft Dialog",
     "'Submit Draft Record' payload sets currentStatus='On Hold'",
     "Draft form filled and submitted",
     "1. Submit a valid draft form; 2. Inspect API payload",
     "POST to /draft-records/<field> with execution_id, form values, currentStatus='On Hold'",
     "","Not Run","High",""),

    ("TC-COR-28","CorrectionsTable","VM CPU(s) Override",
     "For VM sutType, accepted suggestion shows CPU(s) from history (not suggestion raw data)",
     "VM record, suggestion selected, CPU(s) in history",
     "1. Select suggestion in VM record with CPU(s) history",
     "CPU(s) field in SuggestionRow shows history value, displayed correctly",
     "","Not Run","High","getHistoryChangesForField logic; UI override in CorrectionsTable"),

    # ═══════════════════════════════════════════════════════
    # 9. SUGGESTION ROW
    # ═══════════════════════════════════════════════════════
    ("TC-SGR-01","SuggestionRow","Selection",
     "Clicking a suggestion selects it; radio checked, row highlighted",
     "Pending group with suggestions",
     "1. Click any suggestion row",
     "Radio checked; border and background change to SELECTED theme (blue)",
     "","Not Run","Critical",""),

    ("TC-SGR-02","SuggestionRow","Selection",
     "Clicking selected suggestion deselects it",
     "Suggestion selected",
     "1. Click the selected suggestion",
     "Radio unchecked; row returns to white background",
     "","Not Run","High",""),

    ("TC-SGR-03","SuggestionRow","Confidence Tooltip",
     "Hovering over suggestion shows 'Confidence: X%' tooltip",
     "Pending suggestion with score field",
     "1. Hover over a suggestion row",
     "Tooltip appears: 'Confidence: X%'",
     "","Not Run","Medium",""),

    ("TC-SGR-04","SuggestionRow","Internal Fields Filtered",
     "Fields _id, execution_id, snapshot_id, search_key are hidden from suggestion display",
     "Suggestion has these internal fields",
     "1. Open a suggestion row that has internal fields in data",
     "None of _id, execution_id, snapshot_id, search_key appear as visible columns",
     "","Not Run","High","New in v2: internalKeys filter replaces old deduplication approach"),

    ("TC-SGR-05","SuggestionRow","showRadio prop",
     "Draft Record SuggestionRow renders without radio button (showRadio=false)",
     "On Hold group with draft record",
     "1. Observe Draft Record row",
     "No radio button column on the left side of the row",
     "","Not Run","High","New showRadio prop in v2"),

    ("TC-SGR-06","SuggestionRow","EditableField — CPU(s)",
     "CPU(s) field shows edit icon for VM sutType when suggestion selected",
     "VM record; suggestion selected with CPU(s) field",
     "1. Select suggestion in VM record group",
     "Pencil edit icon visible next to the CPU(s) value",
     "","Not Run","High","Field changed from 'coreCount' to 'cpu(s)' in v2"),

    ("TC-SGR-07","SuggestionRow","EditableField — CPU(s)",
     "Inline edit: click pencil, change value, press Enter to save",
     "CPU(s) edit icon visible",
     "1. Click pencil; change value; press Enter",
     "New value shown; edit mode exits; value saved to editedSuggestions state",
     "","Not Run","High",""),

    ("TC-SGR-08","SuggestionRow","EditableField — CPU(s)",
     "Press Escape cancels inline edit and restores original value",
     "Inline edit active",
     "1. Click pencil; change value; press Escape",
     "Original value restored; input closed",
     "","Not Run","High",""),

    ("TC-SGR-09","SuggestionRow","Accepted/OnHold State",
     "Non-pending suggestion rows are read-only (opacity 0.5, cursor not-allowed)",
     "Accepted group",
     "1. Open accepted group, view suggestions",
     "All suggestions dimmed (opacity 0.5); clicking does nothing (cursor: not-allowed)",
     "","Not Run","High",""),

    # ═══════════════════════════════════════════════════════
    # 10. CHOOSE OTHER VALUE DROPDOWN
    # ═══════════════════════════════════════════════════════
    ("TC-CVD-01","ChooseOtherValueDropdown","Visibility",
     "Dropdown visible only for pending groups",
     "Pending field group open",
     "1. Expand a pending group",
     "Autocomplete dropdown 'Choose other <field>:' is present",
     "","Not Run","High",""),

    ("TC-CVD-02","ChooseOtherValueDropdown","Visibility",
     "Dropdown hidden for non-pending groups",
     "Accepted group",
     "1. Expand an accepted group",
     "No dropdown rendered (isPending=false -> returns null)",
     "","Not Run","High",""),

    ("TC-CVD-03","ChooseOtherValueDropdown","Options",
     "Opening dropdown fetches unique masterlist values",
     "Pending group dropdown",
     "1. Click/open the autocomplete",
     "Spinner appears; unique values from /unique-values API loaded as autocomplete options",
     "","Not Run","High",""),

    ("TC-CVD-04","ChooseOtherValueDropdown","Selection",
     "Selecting a value triggers metadata fetch and renders custom SuggestionRow",
     "Dropdown with options loaded",
     "1. Select a value",
     "Dropdown row border turns accent color; 'Custom Value' label appears; SuggestionRow with fetched metadata shown",
     "","Not Run","Critical",""),

    ("TC-CVD-05","ChooseOtherValueDropdown","Clear",
     "Clearing autocomplete clears custom suggestion",
     "Custom value selected",
     "1. Clear the autocomplete (X button)",
     "Custom SuggestionRow disappears; selection deactivated",
     "","Not Run","High",""),

    ("TC-CVD-06","ChooseOtherValueDropdown","Theme Prop",
     "Dropdown receives theme prop and applies activeTheme colors when selected",
     "Custom value selected in pending group",
     "1. Select a custom value",
     "Dropdown row uses SELECTED theme border/background colors",
     "","Not Run","Medium","New theme prop passed in v2"),

    # ═══════════════════════════════════════════════════════
    # 11. REJECT DIALOG (Removed in v2 — split into separate dialogs)
    # ═══════════════════════════════════════════════════════
    ("TC-RJD-01","RejectDialog","Removed",
     "RejectDialog component no longer exists as a separate file",
     "Components directory",
     "1. Check src/components/CorrectionsTableComponents/ directory",
     "RejectDialog.jsx is NOT present; useRejectDialog.js hook is also removed",
     "","Not Run","Critical","BREAKING CHANGE: RejectDialog replaced by L0ConfirmDialog + DraftRecordDialog in v2"),

    ("TC-RJD-02","L0ConfirmDialog","Standalone L0 Flow",
     "L0 flow is now a direct dialog from the main action button bar",
     "Pending group; no suggestion selected",
     "1. Click 'Send to L0'; 2. Observe dialog",
     "Single-step confirmation dialog (no multi-step choose flow); directly confirms L0 send",
     "","Not Run","Critical",""),

    ("TC-RJD-03","DraftRecordDialog","Standalone Draft Flow",
     "Draft flow is now a direct dialog from the main action button bar",
     "Pending group; no suggestion selected",
     "1. Click 'Submit Draft Record'; 2. Observe dialog",
     "Direct Draft Record form dialog (no intermediate choose step)",
     "","Not Run","Critical",""),

    # ═══════════════════════════════════════════════════════
    # 12. ERROR PAGE & LOADER
    # ═══════════════════════════════════════════════════════
    ("TC-ERR-01","ErrorPage","Display",
     "ErrorPage renders with error message",
     "API error during list load",
     "1. Simulate API error; open list page",
     "ErrorPage renders with the error.message and a Retry button",
     "","Not Run","High",""),

    ("TC-ERR-02","ErrorPage","Retry",
     "Retry button on ErrorPage triggers re-fetch",
     "ErrorPage visible",
     "1. Click Retry",
     "Fetch is re-triggered; if API healthy, list loads normally",
     "","Not Run","High",""),

    ("TC-LDR-01","Loader","Spinner",
     "Loader renders centered spinner in all loading states",
     "Any loading condition",
     "1. Trigger loading (page load, record click, etc.)",
     "Centered CircularProgress spinner visible; no layout breakage",
     "","Not Run","Medium",""),

    # ═══════════════════════════════════════════════════════
    # 13. RESPONSIVE DESIGN
    # ═══════════════════════════════════════════════════════
    ("TC-RES-01","Responsive","Mobile (<768px)",
     "Search bar adapts to 90% width on mobile",
     "Mobile viewport",
     "1. Resize to 375px; observe search field",
     "Width is ~90% of viewport; no overflow",
     "","Not Run","High",""),

    ("TC-RES-02","Responsive","Mobile (<768px)",
     "Filter buttons wrap on narrow viewports",
     "Mobile viewport with filters shown",
     "1. Open /active on mobile",
     "Filter buttons wrap to next line; no horizontal scrollbar",
     "","Not Run","High",""),

    ("TC-RES-03","Responsive","Tablet (768-1200px)",
     "Record cards at 65vw readable on tablet",
     "Tablet viewport",
     "1. Open list page at 768px wide",
     "Cards at 65vw; text not clipped; cards centered",
     "","Not Run","Medium",""),

    # ═══════════════════════════════════════════════════════
    # 14. CROSS-BROWSER
    # ═══════════════════════════════════════════════════════
    ("TC-CBR-01","Cross-Browser","Chrome",
     "App renders and all flows work on Chrome (latest)",
     "Chrome latest installed",
     "1. Open app in Chrome; test all main flows",
     "No rendering issues; all dialogs and cards render correctly",
     "","Not Run","Critical",""),

    ("TC-CBR-02","Cross-Browser","Firefox",
     "App works correctly on Firefox (latest)",
     "Firefox installed",
     "1. Open in Firefox; test main flows",
     "No issues",
     "","Not Run","High",""),

    ("TC-CBR-03","Cross-Browser","Edge",
     "App works on Microsoft Edge",
     "Edge installed",
     "1. Open app in Edge",
     "No rendering or functionality issues",
     "","Not Run","Medium",""),

    # ═══════════════════════════════════════════════════════
    # 15. PERFORMANCE
    # ═══════════════════════════════════════════════════════
    ("TC-PER-01","Performance","Load Time",
     "List page shows first 50 records within 3 seconds",
     "Fresh page load; backend running",
     "1. Open /active; time to first record render",
     "Records visible within 3 seconds",
     "","Not Run","High",""),

    ("TC-PER-02","Performance","Search Debounce",
     "Results update within 1 second after 500ms debounce",
     "List page loaded",
     "1. Type in search; wait for debounce",
     "API fires once; results update within 1s of debounce completing",
     "","Not Run","High",""),

    ("TC-PER-03","Performance","Memory / Abort",
     "Navigating between list pages aborts stale in-flight requests",
     "Rapidly switching routes",
     "1. Quickly switch between /active and /completed several times",
     "Only current page results shown; no ghost data; no console AbortError crashes",
     "","Not Run","High",""),

    ("TC-PER-04","Performance","Animation Smoothness",
     "Card hover and filter button transitions are smooth (no jank)",
     "Any list page",
     "1. Rapidly hover over cards and filter buttons",
     "Transitions feel smooth; no flickering or layout shift",
     "","Not Run","Medium",""),

    # ═══════════════════════════════════════════════════════
    # 16. ACCESSIBILITY
    # ═══════════════════════════════════════════════════════
    ("TC-ACC-01","Accessibility","Keyboard",
     "Search field focusable via Tab key",
     "Any list page",
     "1. Tab to search; type query",
     "Focus indicator visible; search works via keyboard",
     "","Not Run","High",""),

    ("TC-ACC-02","Accessibility","Keyboard",
     "Filter buttons activatable via keyboard (Space/Enter)",
     "List page with filters",
     "1. Tab to filter button; press Space or Enter",
     "Filter activates; list updates",
     "","Not Run","High",""),

    ("TC-ACC-03","Accessibility","Contrast",
     "All text meets WCAG AA contrast ratio (>=4.5:1)",
     "Any page",
     "1. Run Lighthouse accessibility audit",
     "All text passes AA contrast",
     "","Not Run","Medium",""),

    ("TC-ACC-04","Accessibility","ARIA",
     "Dialogs have ARIA role and focus trap",
     "Any dialog open",
     "1. Open Accept, L0, or Draft dialog; check ARIA",
     "role='dialog'; focus trapped inside dialog while open",
     "","Not Run","Medium",""),
]

# ─────────────────────────────────────────────────────────────────────────────
# WORKBOOK CONSTRUCTION
# ─────────────────────────────────────────────────────────────────────────────
wb = openpyxl.Workbook()
wb.remove(wb.active)

# ── Cover Page ───────────────────────────────────────────────────────────────
ws_cov = wb.create_sheet("Cover Page")
ws_cov.sheet_view.showGridLines = False
ws_cov.column_dimensions["A"].width = 5
ws_cov.column_dimensions["B"].width = 55
ws_cov.column_dimensions["C"].width = 35

def cw(r, c, v, bg=None, fg="FF000000", bold=False, sz=11, ha="left", wrap=False, italic=False):
    cell = ws_cov.cell(row=r, column=c, value=v)
    if bg: cell.fill = fill(bg)
    cell.font = font(size=sz, bold=bold, color=fg, italic=italic)
    cell.alignment = align(ha, wrap=wrap)
    return cell

ws_cov.merge_cells("A1:C1"); ws_cov.row_dimensions[1].height = 8
ws_cov.merge_cells("A2:C2")
cw(2,1,"AMD Data Hygiene — QA Test Report v2",bg=C["navy"],fg=C["white"],bold=True,sz=22,ha="center")
ws_cov.row_dimensions[2].height = 50
ws_cov.merge_cells("A3:C3")
cw(3,1,"Frontend & UI Test Cases — post new-master-list merge",bg=C["navy_light"],fg=C["white"],sz=13,ha="center",italic=True)
ws_cov.row_dimensions[3].height = 28
ws_cov.merge_cells("A4:C4"); ws_cov.row_dimensions[4].height = 12

info = [
    ("Project",      "AMD Data Hygiene"),
    ("Version",      "2.0 (new-master-list)"),
    ("Report Date",  str(date.today().strftime("%B %d, %Y"))),
    ("Prepared By",  "QA Team"),
    ("Branch",       "quality-analysis / new-master-list"),
    ("Framework",    "React + Vite + Material UI"),
    ("Test Scope",   "UI / Frontend — All Components & Pages (v2 changes)"),
    ("Total TCs",    str(len(TESTS))),
]
for i,(k,v) in enumerate(info):
    r = 5+i; ws_cov.row_dimensions[r].height = 24
    cw(r,1,"",bg=C["off_white"])
    cw(r,2,k,bg=C["gray_bg"],bold=True,fg=C["navy"],sz=11)
    cw(r,3,v,bg=C["off_white"],fg="FF1E293B",sz=11)

ws_cov.merge_cells("A14:C14"); ws_cov.row_dimensions[14].height = 16
ws_cov.merge_cells("A15:C15")
cw(15,1,"STATUS LEGEND",bg=C["navy"],fg=C["white"],bold=True,sz=12,ha="center")
ws_cov.row_dimensions[15].height = 26
for j,(s,bg,fg_c) in enumerate([
    ("Not Run",C["notrun_bg"],C["notrun_fg"]),
    ("Pass",   C["pass_bg"],  C["pass_fg"]),
    ("Fail",   C["fail_bg"],  C["fail_fg"]),
    ("Blocked",C["blocked_bg"],C["blocked_fg"]),
    ("N/A",    C["na_bg"],    C["na_fg"]),
]):
    r = 16+j; ws_cov.row_dimensions[r].height = 22
    ws_cov.merge_cells(f"B{r}:C{r}")
    cc = ws_cov.cell(row=r,column=2,value=s)
    cc.fill=fill(bg); cc.font=font(bold=True,color=fg_c,size=11); cc.alignment=align("center")

ws_cov.merge_cells("A22:C22"); ws_cov.row_dimensions[22].height = 16
ws_cov.merge_cells("A23:C23")
cw(23,1,"PRIORITY LEGEND",bg=C["navy"],fg=C["white"],bold=True,sz=12,ha="center")
ws_cov.row_dimensions[23].height = 26
for j,(p,c_) in enumerate([
    ("Critical","FF991B1B"),("High","FF9A3412"),("Medium","FF854D0E"),("Low","FF166534")
]):
    r=24+j; ws_cov.row_dimensions[r].height=22
    ws_cov.merge_cells(f"B{r}:C{r}")
    cc=ws_cov.cell(row=r,column=2,value=p)
    cc.fill=fill(C["off_white"]); cc.font=font(bold=True,color=c_,size=11); cc.alignment=align("center")

ws_cov.merge_cells("A29:C29"); ws_cov.row_dimensions[29].height = 16
ws_cov.merge_cells("A30:C30")
cw(30,1,"KEY CHANGES vs v1",bg=C["navy"],fg=C["white"],bold=True,sz=12,ha="center")
ws_cov.row_dimensions[30].height = 26
changes_v2 = [
    "Status filter 'Rejected' renamed to 'L0 Data' (API value unchanged)",
    "Card Status display: rejected records now show 'L0 Data'",
    "Card background now always white; differentiation via left border only",
    "On Hold color scheme changed from purple to yellow throughout",
    "RejectDialog removed — replaced by L0ConfirmDialog + DraftRecordDialog",
    "New action button: 'Submit Draft Record' (yellow) + 'Send to L0' (red)",
    "Draft dialog: dynamic fields, integer type support, history pre-fill",
    "Pending chip added to CorrectionsTable group header",
    "Groups auto-expand only for pending; others collapsed by default",
    "Draft Record section (read-only SuggestionRow) shown for On Hold groups",
    "SuggestionRow: internal keys filtered; showRadio prop added",
    "Editable field changed from 'coreCount' to 'cpu(s)'",
    "Route keys added to prevent stale state between list pages",
    "/completed/:id route removed",
]
for j, txt in enumerate(changes_v2):
    r = 31+j; ws_cov.row_dimensions[r].height = 20
    ws_cov.merge_cells(f"B{r}:C{r}")
    cc = ws_cov.cell(row=r, column=2, value=f"  • {txt}")
    cc.fill = fill(C["off_white"] if j%2==0 else C["white"])
    cc.font = font(size=10, color="FF1E293B"); cc.alignment = align("left", wrap=True)

# ── Test Summary ─────────────────────────────────────────────────────────────
ws_sum = wb.create_sheet("Test Summary")
ws_sum.sheet_view.showGridLines = False

MODULES = list(OrderedDict.fromkeys(tc[1] for tc in TESTS))

counts = defaultdict(lambda: defaultdict(int))
total_counts = defaultdict(int)
for tc in TESTS:
    counts[tc[1]][tc[8]] += 1
    total_counts[tc[8]] += 1

ws_sum.merge_cells("A1:H1")
t = ws_sum.cell(row=1,column=1,value="Test Summary Dashboard — v2")
t.fill=fill(C["navy"]); t.font=font(size=16,bold=True,color=C["white"])
t.alignment=align("center"); ws_sum.row_dimensions[1].height=40

headers = ["Module","Not Run","Pass","Fail","Blocked","N/A","Total","% Pass"]
col_widths = [34,12,10,10,12,10,10,12]
for ci,(h,w) in enumerate(zip(headers,col_widths),1):
    ws_sum.column_dimensions[get_column_letter(ci)].width = w
    ch(ws_sum, 2, ci, h)

for ri, mod in enumerate(MODULES, 3):
    nr  = counts[mod].get("Not Run",0); pas = counts[mod].get("Pass",0)
    fai = counts[mod].get("Fail",0);   blo = counts[mod].get("Blocked",0)
    na_ = counts[mod].get("N/A",0);    tot = nr+pas+fai+blo+na_
    pct = f"{round(pas/tot*100)}%" if tot else "-"
    bg  = C["off_white"] if ri%2==0 else C["white"]
    wc(ws_sum,ri,1,mod,     bg=bg,bold=True,fg=C["navy"])
    wc(ws_sum,ri,2,nr,      bg=C["notrun_bg"],fg=C["notrun_fg"],h_align="center")
    wc(ws_sum,ri,3,pas,     bg=C["pass_bg"],  fg=C["pass_fg"],  h_align="center")
    wc(ws_sum,ri,4,fai,     bg=C["fail_bg"],  fg=C["fail_fg"],  h_align="center")
    wc(ws_sum,ri,5,blo,     bg=C["blocked_bg"],fg=C["blocked_fg"],h_align="center")
    wc(ws_sum,ri,6,na_,     bg=C["na_bg"],    fg=C["na_fg"],    h_align="center")
    wc(ws_sum,ri,7,tot,     bg=bg,bold=True,fg=C["navy"],h_align="center")
    wc(ws_sum,ri,8,pct,     bg=bg,h_align="center",fg=C["gray_text"])
    ws_sum.row_dimensions[ri].height = 22

tr = len(MODULES)+3
wc(ws_sum,tr,1,"TOTAL",bg=C["navy"],fg=C["white"],bold=True)
for ci,key in enumerate(["Not Run","Pass","Fail","Blocked","N/A"],2):
    bg_,fg_ = STATUS_STYLE[key]
    wc(ws_sum,tr,ci,total_counts[key],bg=bg_,fg=fg_,bold=True,h_align="center")
wc(ws_sum,tr,7,len(TESTS),bg=C["navy"],fg=C["white"],bold=True,h_align="center")
ws_sum.row_dimensions[tr].height=26

# ── All Test Cases Sheet ──────────────────────────────────────────────────────
ws_tc = wb.create_sheet("Test Cases")
ws_tc.sheet_view.showGridLines = False
ws_tc.freeze_panes = "A3"

TC_COLS = [
    ("TC ID",10),("Module",24),("Component",28),("Test Case",52),
    ("Pre-conditions",30),("Steps",52),("Expected Result",52),
    ("Actual Result",38),("Status",13),("Priority",13),("Remarks",32),
]
for i,(h,w) in enumerate(TC_COLS):
    ws_tc.column_dimensions[get_column_letter(i+1)].width = w

ws_tc.merge_cells(f"A1:{get_column_letter(len(TC_COLS))}1")
t2 = ws_tc.cell(row=1,column=1,value="AMD Data Hygiene v2 — UI/Frontend Test Cases")
t2.fill=fill(C["navy"]); t2.font=font(size=14,bold=True,color=C["white"])
t2.alignment=align("center"); ws_tc.row_dimensions[1].height=36

for i,(h,_) in enumerate(TC_COLS): ch(ws_tc,2,i+1,h)
ws_tc.row_dimensions[2].height=24

last_module = None
data_row = 3
for tc in TESTS:
    tc_id,module,comp,desc,pre,steps,exp,actual,status,pri,remarks = tc
    if module != last_module:
        ws_tc.merge_cells(start_row=data_row,start_column=1,end_row=data_row,end_column=len(TC_COLS))
        sep = ws_tc.cell(row=data_row,column=1,value=f"  | {module.upper()}")
        sep.fill=fill(C["navy_light"]); sep.font=font(size=11,bold=True,color=C["white"])
        sep.alignment=align("left"); ws_tc.row_dimensions[data_row].height=20
        data_row+=1; last_module=module
    bg = C["off_white"] if data_row%2==0 else C["white"]
    wc(ws_tc,data_row,1,tc_id,  bg=C["blue_light"],bold=True,fg=C["navy"],size=10)
    wc(ws_tc,data_row,2,module, bg=bg,fg=C["navy"],bold=True,size=10)
    wc(ws_tc,data_row,3,comp,   bg=bg,size=10)
    wc(ws_tc,data_row,4,desc,   bg=bg,size=10,wrap=True)
    wc(ws_tc,data_row,5,pre,    bg=bg,size=10,wrap=True,fg=C["gray_text"],italic=True)
    wc(ws_tc,data_row,6,steps,  bg=bg,size=10,wrap=True)
    wc(ws_tc,data_row,7,exp,    bg=bg,size=10,wrap=True,fg="FF065F46")
    wc(ws_tc,data_row,8,actual, bg=bg,size=10,wrap=True,fg=C["gray_text"],italic=True)
    apply_status(ws_tc.cell(row=data_row,column=9,value=status), status)
    apply_priority(ws_tc.cell(row=data_row,column=10,value=pri), pri)
    wc(ws_tc,data_row,11,remarks,bg=bg,size=10,italic=True,fg=C["gray_text"])
    ws_tc.row_dimensions[data_row].height=60
    data_row+=1

# ── Per-Module Sheets ─────────────────────────────────────────────────────────
by_module = OrderedDict()
for tc in TESTS: by_module.setdefault(tc[1],[]).append(tc)

for mod, tcs in by_module.items():
    safe = mod.replace("/","-").replace(" ","_")[:28]
    ws = wb.create_sheet(safe)
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A3"
    cols = [("TC ID",10),("Component",28),("Test Case",50),("Steps",50),
            ("Expected Result",50),("Status",12),("Priority",12),("Remarks",30)]
    for i,(h,w) in enumerate(cols):
        ws.column_dimensions[get_column_letter(i+1)].width = w
    ws.merge_cells(f"A1:{get_column_letter(len(cols))}1")
    t3 = ws.cell(row=1,column=1,value=f"{mod} — Test Cases")
    t3.fill=fill(C["navy"]); t3.font=font(size=13,bold=True,color=C["white"])
    t3.alignment=align("center"); ws.row_dimensions[1].height=34
    for i,(h,_) in enumerate(cols): ch(ws,2,i+1,h)
    ws.row_dimensions[2].height=22
    for dr, tc in enumerate(tcs,3):
        tc_id,_,comp,desc,_,steps,exp,_,status,pri,rem = tc
        bg = C["off_white"] if dr%2==0 else C["white"]
        wc(ws,dr,1,tc_id,  bg=C["blue_light"],bold=True,fg=C["navy"],size=10)
        wc(ws,dr,2,comp,   bg=bg,size=10)
        wc(ws,dr,3,desc,   bg=bg,size=10,wrap=True)
        wc(ws,dr,4,steps,  bg=bg,size=10,wrap=True)
        wc(ws,dr,5,exp,    bg=bg,size=10,wrap=True,fg="FF065F46")
        apply_status(ws.cell(row=dr,column=6,value=status), status)
        apply_priority(ws.cell(row=dr,column=7,value=pri), pri)
        wc(ws,dr,8,rem,    bg=bg,size=10,italic=True,fg=C["gray_text"])
        ws.row_dimensions[dr].height=55

# ── Bug Tracker ───────────────────────────────────────────────────────────────
ws_bug = wb.create_sheet("Bug Tracker")
ws_bug.sheet_view.showGridLines = False
ws_bug.freeze_panes = "A3"
bug_cols = [
    ("Bug ID",10),("TC ID",12),("Module",22),("Component",26),("Bug Title",48),
    ("Description",48),("Severity",14),("Priority",12),("Status",16),
    ("Reported By",20),("Reported Date",18),("Fixed By",20),("Fixed Date",18),("Remarks",28),
]
for i,(h,w) in enumerate(bug_cols):
    ws_bug.column_dimensions[get_column_letter(i+1)].width = w
ws_bug.merge_cells(f"A1:{get_column_letter(len(bug_cols))}1")
bt = ws_bug.cell(row=1,column=1,value="Bug Tracker")
bt.fill=fill(C["navy"]); bt.font=font(size=14,bold=True,color=C["white"])
bt.alignment=align("center"); ws_bug.row_dimensions[1].height=36
for i,(h,_) in enumerate(bug_cols): ch(ws_bug,2,i+1,h)
ws_bug.row_dimensions[2].height=22
for r in range(3,23):
    bg = C["off_white"] if r%2==0 else C["white"]
    for c in range(1,len(bug_cols)+1):
        cell = ws_bug.cell(row=r,column=c)
        cell.fill=fill(bg); cell.alignment=align(wrap=True); cell.border=bdr()
    ws_bug.row_dimensions[r].height=28
wc(ws_bug,3,7,"[Blocker/Critical/Major/Minor/Trivial]",bg=C["off_white"],fg=C["gray_text"],size=9,italic=True,wrap=True)
wc(ws_bug,3,9,"[Open/In Progress/Fixed/Verified/Closed]",bg=C["off_white"],fg=C["gray_text"],size=9,italic=True,wrap=True)

# ── Execution Log ─────────────────────────────────────────────────────────────
ws_log = wb.create_sheet("Execution Log")
ws_log.sheet_view.showGridLines = False
log_cols = [
    ("Run #",8),("Date",16),("Tester",20),("Environment",20),
    ("Build/Branch",22),("TCs Run",10),("Passed",10),("Failed",10),
    ("Blocked",10),("Not Run",10),("Pass %",12),("Notes",50),
]
for i,(h,w) in enumerate(log_cols):
    ws_log.column_dimensions[get_column_letter(i+1)].width = w
ws_log.merge_cells(f"A1:{get_column_letter(len(log_cols))}1")
lt = ws_log.cell(row=1,column=1,value="Test Execution Log")
lt.fill=fill(C["navy"]); lt.font=font(size=14,bold=True,color=C["white"])
lt.alignment=align("center"); ws_log.row_dimensions[1].height=36
for i,(h,_) in enumerate(log_cols): ch(ws_log,2,i+1,h)
ws_log.row_dimensions[2].height=22
for r in range(3,18):
    bg = C["off_white"] if r%2==0 else C["white"]
    for c in range(1,len(log_cols)+1):
        cell=ws_log.cell(row=r,column=c)
        cell.fill=fill(bg); cell.alignment=align(wrap=True); cell.border=bdr()
    ws_log.row_dimensions[r].height=28
run1 = ["1",str(date.today().strftime("%Y-%m-%d")),"","Local Dev",
        "quality-analysis/new-master-list",len(TESTS),0,0,0,len(TESTS),"0%","Initial run v2 — all TCs pending"]
for ci,val in enumerate(run1,1): wc(ws_log,3,ci,val,bg=C["off_white"],size=10)

# ── Sheet order ───────────────────────────────────────────────────────────────
fixed_sheets = ["Cover Page","Test Summary","Test Cases","Bug Tracker","Execution Log"]
module_sheets = [s for s in wb.sheetnames if s not in fixed_sheets]
for i,name in enumerate(fixed_sheets+module_sheets):
    wb.move_sheet(name, offset=i-wb.sheetnames.index(name))

# ── Save ──────────────────────────────────────────────────────────────────────
OUT = r"d:\DH\Data-Hygiene\QA_Report_DataHygiene_UI_v2.xlsx"
wb.save(OUT)
print(f"[OK] QA report v2 saved -> {OUT}")
print(f"     Total test cases: {len(TESTS)}")
print(f"     Sheets: {wb.sheetnames}")

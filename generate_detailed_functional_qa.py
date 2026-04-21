"""
Detailed Functional QA Report Generator — AMD Data Hygiene Frontend
Includes explicit Positive and Negative test cases with granular interaction validation.
"""

import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Palette
C = {
    "navy": "FF17233A",
    "white": "FFFFFFFF",
    "gray_bg": "FFF1F5F9",
    "notrun_bg": "FFF5F0FF", "notrun_fg": "FF5B21B6",
    "pos_bg": "FFD1FAE5", "pos_fg": "FF065F46",       # Green for positive
    "neg_bg": "FFFEE2E2", "neg_fg": "FF991B1B",       # Red for negative
}

def fill(h): return PatternFill("solid", fgColor=h)
def font(bold=False, color="FF000000", size=10): return Font(name="Calibri", size=size, bold=bold, color=color)
def align(h="left", wrap=True): return Alignment(horizontal=h, vertical="center", wrap_text=wrap)
def bdr(): return Border(left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin"))

# [TC_ID, Type, Workflow Area, Component, Scenario / Interaction, Expected System Behavior, Status, Priority]
TESTS = [
    # ── SEARCH BAR ────────────────────────────────────────────────
    ("F-SRC-01", "Positive", "Search Bar", "Input Debounce", 
     "Type 'test', pause 200ms, type 'ing', pause 600ms.", 
     "API call fires once, exactly ~500ms after 'ing'. Spinner visible during flight.", "Not Run", "Critical"),
    ("F-SRC-02", "Positive", "Search Bar", "Clear & Re-fetch", 
     "Delete all text (Backspace).", 
     "Debounce triggers. API called with empty search param. List is refreshed to base state.", "Not Run", "High"),
    ("F-SRC-N1", "Negative", "Search Bar", "Cross-Site Scripting (XSS)", 
     "Type `<script>alert('xss')</script>`.", 
     "Search executes. React escapes HTML entities. Returns empty state. No script executes.", "Not Run", "Critical"),
    ("F-SRC-N2", "Negative", "Search Bar", "Extreme Length Input", 
     "Paste a contiguous 10,000 character string into the search input.", 
     "App does not crash or freeze. Debounce fires safely. API handles or truncates gracefully.", "Not Run", "Medium"),

    # ── FILTER BUTTON LOGIC ──────────────────────────────────────────────────
    ("F-FLT-01", "Positive", "Filter Buttons", "Single Toggle", 
     "Click 'Accepted' filter -> verify list -> Click 'Accepted' again.", 
     "Active state works. Second click returns to default button UI and default route fetch.", "Not Run", "Critical"),
    ("F-FLT-N1", "Negative", "Filter Buttons", "Rapid Spam Clicking", 
     "Click the 'On Hold' filter rapidly 20 times in 2 seconds.", 
     "App restricts fetch thrashing via abort controllers or internal lockout. Only the final state fetches, avoiding UI race conditions.", "Not Run", "High"),

    # ── SCROLLING & PAGINATION ─────────────────────────────────────────
    ("F-PAG-01", "Positive", "Pagination", "Infinite Scroll Trigger", 
     "Slowly scroll down a list of >50 items.", 
     "loadMore function triggers automatically. Temporary bottom loader appears. New records append.", "Not Run", "Critical"),
    ("F-PAG-N1", "Negative", "Pagination", "Out of Bounds Simulation", 
     "Mock API response `page: 5, totalPages: 2`.", 
     "Client logic recognizes page > totalPages and gracefully stops attempting loadMoreRows.", "Not Run", "High"),

    # ── CORRECTIONS TABLE EXPANSION ──────────────────────────────────────────
    ("F-CXP-01", "Positive", "Corrections Expansion", "Default State Logic", 
     "Open a record card mixing 'Pending', 'Accepted', and 'On Hold'.", 
     "Pending auto-expands. Accepted and On Hold auto-collapse.", "Not Run", "Critical"),
    ("F-CXP-N1", "Negative", "Corrections Expansion", "Malformed Status", 
     "Mock record where currentStatus is `null` or unmapped string `UNKNOWN_ERROR`.", 
     "Fallback logic defaults to treating it as 'pending' (expanded) with a warning, rather than crashing the component tree.", "Not Run", "Medium"),

    # ── SUGGESTION ROW INTERACTIONS ─────────────────────────────────────────
    ("F-SGR-01", "Positive", "Suggestion Selection", "Mutual Exclusivity", 
     "Select Suggestion A -> Select Custom Dropdown Value.", 
     "Selecting custom value immediately deselects Suggestion A.", "Not Run", "Critical"),
    ("F-SGR-N1", "Negative", "Suggestion Selection", "Double Selection Exploit", 
     "Attempt to use Tab/Spacebar to select multiple radio boxes via keyboard concurrently.", 
     "State handlers enforce strict singular selection per invalid_field group under all circumstances.", "Not Run", "High"),

    # ── VM CPU INLINE EDITING ────────────────────────────────────────────────
    ("F-CPU-01", "Positive", "Inline Edit", "Initiation & Save (VM only)", 
     "Click Pencil on CPU(s) -> type new number -> Press Enter.", 
     "Saves to local state (editedSuggestions), closes input, updates display.", "Not Run", "High"),
    ("F-CPU-N1", "Negative", "Inline Edit", "Non-numeric Input", 
     "Click Pencil -> Type 'abcdef' -> Press Enter.", 
     "UI blocks save, alerts user, or forces sanitization to numeric values only before finalizing the edit state.", "Not Run", "High"),

    # ── CUSTOM VALUE DROPDOWN ────────────────────────────────────────────────
    ("F-CVD-01", "Positive", "Custom Dropdown", "Metadata Fetch Flow", 
     "Open dropdown -> Select an unlisted value -> Watch UI.", 
     "Fetch yields metadata -> dynamic SuggestionRow appears labeled 'Custom Value'.", "Not Run", "High"),
    ("F-CVD-N1", "Negative", "Custom Dropdown", "Metadata Fetch Timeout", 
     "Select Custom Value -> force API timeout or 404.", 
     "Spinner halts. Error snackbar appears. Autocomplete selection is reverted/cleared.", "Not Run", "Critical"),

    # ── SUBMISSION WORKFLOWS & API FAILURES ───────────────────────────────
    ("F-FLW-01", "Positive", "L0 Flow", "Submission", 
     "Click 'Send to L0' -> Click 'Yes, Confirm'.", 
     "PUT /reject-record triggers. Dialog closes, data refreshed, Green snackbar shown.", "Not Run", "Critical"),
    ("F-FLW-N1", "Negative", "Submit Flows", "500 Server Error Response", 
     "Click 'Accept' -> Mock 500 Internal Server error.", 
     "Accepting state un-toggles (buttons unlock). Dialog remains open or closes safely. Red Error Snackbar appears with specific server error text.", "Not Run", "Critical"),
    ("F-LCK-N1", "Negative", "Submit Flows", "Mid-flight Cancel Lockout", 
     "Click 'Accept' -> API takes 3 seconds -> rapidly click Dialog 'Close (X)' during flight.", 
     "Dialog refuses to close. Redundant API calls are blocked. System maintains data integrity.", "Not Run", "High"),

    # ── DRAFT RECORD FLOW & VALIDATION ───────────────────────────────────────
    ("F-DRF-01", "Positive", "Draft Flow", "Dynamic Field Loading", 
     "Click 'Submit Draft Record'.", 
     "Fields load dynamically from /draft-records/fields. History pre-fills accurately.", "Not Run", "Critical"),
    ("F-DRF-N1", "Negative", "Draft Flow", "Float/Decimal Input Rejection", 
     "Type '14.5' into an 'integer' marked dynamic field -> Submit.", 
     "Submission HALTS locally. Error Snackbar appears specifying whole-numbers are required.", "Not Run", "Critical"),
    ("F-DRF-N2", "Negative", "Draft Flow", "Negative Number Bounds", 
     "Type '-5' into an 'integer' marked dynamic field -> Submit.", 
     "Input `min=0` constraint enforced. Form submission blocked locally with appropriate error.", "Not Run", "High"),
    ("F-DRF-N3", "Negative", "Draft Flow", "Blank Field Violation", 
     "Bypass UI disable -> force a blank string representation in payload.", 
     "Locally restricted: Submit button disables if field value `.trim() === ''`.", "Not Run", "High"),

    # ── REACT VIRTUALIZATION ────────────────────────────────────────────────
    ("F-VRT-01", "Positive", "Virtualization", "Dynamic Row Heights", 
     "Scroll through a list mixing 'Accepted' and 'Pending' records.", 
     "Accepted map to 100px. Pending map to 160px. No card clipping/overlapping.", "Not Run", "Critical"),
    ("F-VRT-N1", "Negative", "Virtualization", "Resize Throttling", 
     "Violently resize browser window horizontally continually for 5 seconds.", 
     "WindowScroller+AutoSizer handles resize limits. App does not encounter 'ResizeObserver loop limit exceeded' crashes. Stays performant.", "Not Run", "Medium"),
]

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Detailed Functional QA"
ws.sheet_view.showGridLines = False

# Header
ws.merge_cells("A1:H1")
header = ws.cell(row=1, column=1, value="Granular Interactive Workflows — Positive & Negative Tests")
header.fill = fill(C["navy"]); header.font = font(bold=True, size=14, color=C["white"]); header.alignment = align("center")
ws.row_dimensions[1].height = 30

# Cols
cols = ["TC ID", "Test Type", "Workflow Area", "Component / Action", "Detailed Scenario / Interaction", "Strict System Behavior & Validation", "Status", "Priority"]
widths = [11, 12, 20, 23, 40, 50, 12, 12]

for i, h in enumerate(cols, 1):
    cell = ws.cell(row=2, column=i, value=h)
    cell.fill = fill(C["navy"]); cell.font = font(bold=True, color=C["white"]); cell.alignment = align("center")
    ws.column_dimensions[get_column_letter(i)].width = widths[i-1]

# Rows
for r, (tcid, ttype, area, comp, scen, expt, stat, pri) in enumerate(TESTS, 3):
    bg = C["white"] if r % 2 == 0 else C["gray_bg"]
    
    # ID
    ws.cell(row=r, column=1, value=tcid).fill = fill(bg)
    ws.cell(row=r, column=1).border = bdr()
    
    # Test Type
    type_cell = ws.cell(row=r, column=2, value=ttype)
    type_cell.border = bdr(); type_cell.alignment = align("center")
    if ttype == "Positive":
        type_cell.fill = fill(C["pos_bg"]); type_cell.font = font(color=C["pos_fg"], bold=True)
    else:
        type_cell.fill = fill(C["neg_bg"]); type_cell.font = font(color=C["neg_fg"], bold=True)
        
    # Basic cols
    for c, val in enumerate([area, comp, scen, expt], 3):
        cell = ws.cell(row=r, column=c, value=val)
        cell.fill = fill(bg); cell.alignment = align(); cell.font = font(); cell.border = bdr()
    
    # Status
    s_cell = ws.cell(row=r, column=7, value=stat)
    s_cell.fill = fill(C["notrun_bg"]); s_cell.font = font(bold=True, color=C["notrun_fg"]); s_cell.alignment = align("center"); s_cell.border = bdr()
    
    # Priority
    p_cell = ws.cell(row=r, column=8, value=pri)
    p_cell.fill = fill(bg); p_cell.alignment = align("center"); p_cell.border = bdr()
    p_color = {"Critical": "FF991B1B", "High": "FF9A3412", "Medium": "FF854D0E", "Low": "FF166534"}.get(pri, "FF000000")
    p_cell.font = font(bold=True, color=p_color)

OUTPUT_PATH = r"d:\DH\Data-Hygiene\QA_Pos_Neg_Functional_Report.xlsx"
wb.save(OUTPUT_PATH)
print(f"Report saved to: {OUTPUT_PATH}")

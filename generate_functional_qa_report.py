"""
Functional QA Report Generator — AMD Data Hygiene Frontend
Focuses strictly on INTERACTABLE functionalities and logic flows.
"""

import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import date

# Palette for formatting
C = {
    "navy": "FF17233A",
    "white": "FFFFFFFF",
    "gray_bg": "FFF1F5F9",
    "pass_bg": "FFD1FAE5", "pass_fg": "FF065F46",
    "fail_bg": "FFFEE2E2", "fail_fg": "FF991B1B",
    "notrun_bg": "FFF5F0FF", "notrun_fg": "FF5B21B6",
}

def fill(h): return PatternFill("solid", fgColor=h)
def font(bold=False, color="FF000000", size=11): return Font(name="Calibri", size=size, bold=bold, color=color)
def align(h="left", wrap=True): return Alignment(horizontal=h, vertical="center", wrap_text=wrap)
def bdr(): return Border(left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin"))

# Functional Test Cases
# [TC_ID, Area, Component, Interaction, Expected Result, Status, Priority]
FUNCTIONAL_TESTS = [
    # ── SEARCH & FILTER ───────────────────────────────────────────────────
    ("F-SF-01", "Search & Filter", "Search Bar", 
     "Type text and wait 500ms (Debounce)", 
     "API call should only trigger once, 500ms after the user stops typing. Spinner should show during wait.", 
     "Not Run", "Critical"),
    ("F-SF-02", "Search & Filter", "Search Bar", 
     "Clear search text via backspace", 
     "List should automatically refresh to show all records (or filtered set) after debounce.", 
     "Not Run", "High"),
    ("F-SF-03", "Search & Filter", "Status Filters", 
     "Click 'Pending' filter on Landing Page", 
     "API request should include 'status=pending'. List should update to only show pending records.", 
     "Not Run", "Critical"),
    ("F-SF-04", "Search & Filter", "Status Filters", 
     "Click an already active filter button", 
     "Filter should toggle OFF, and the list should fetch all records (clearing the status param).", 
     "Not Run", "High"),
    ("F-SF-05", "Search & Filter", "Age Filters", 
     "Click '< 3 Days' on Active List", 
     "API request should combine default status (pending) with age=green. List updates.", 
     "Not Run", "High"),

    # ── NAVIGATION ────────────────────────────────────────────────────────
    ("F-NAV-01", "Navigation", "Navbar", 
     "Click 'Completed List' while on 'Active List'", 
     "Route should change, and page state (search/filters) should reset completely due to key change.", 
     "Not Run", "High"),
    ("F-NAV-02", "Navigation", "Record Card", 
     "Click anywhere on a record card", 
     "User should be navigated to the Details Page (/:id) for that specific record.", 
     "Not Run", "Critical"),

    # ── LIST INTERACTION ──────────────────────────────────────────────────
    ("F-LIST-01", "List Loading", "Infinite Scroll", 
     "Scroll to the bottom of a long list", 
     "Next page of records should fetch and append automatically. No manual click required.", 
     "Not Run", "Critical"),
    ("F-LIST-02", "List Loading", "Error Retry", 
     "Click 'Retry' on Error Page after a failed fetch", 
     "The app should re-attempt the API call and reload the list if the server response is successful.", 
     "Not Run", "High"),

    # ── CORRECTIONS TABLE LOGIC ───────────────────────────────────────────
    ("F-COR-01", "Corrections", "Group Toggle", 
     "Click a group header (e.g., 'instanceType')", 
     "Group should expand to reveal suggestions if collapsed, or collapse if expanded.", 
     "Not Run", "Medium"),
    ("F-COR-02", "Corrections", "Suggestion Selection", 
     "Click a suggestion row", 
     "Radio button should be checked, row highlighted, and the 'Accept' button should ENABLE.", 
     "Not Run", "Critical"),
    ("F-COR-03", "Corrections", "Suggestion Deselection", 
     "Click a currently selected suggestion row", 
     "Radio should uncheck, and the 'Accept' button should DISABLE.", 
     "Not Run", "High"),
    ("F-COR-04", "Corrections", "Inline Edit (VM)", 
     "Click pencil icon on CPU(s) field in a VM record suggestion", 
     "Enter edit mode. Changing numeric value and pressing Enter should update the local selection state.", 
     "Not Run", "High"),
    ("F-COR-05", "Corrections", "Custom Dropdown", 
     "Type and select a value in 'Choose other <field>' Autocomplete", 
     "App should fetch metadata for selection, render it as a 'Custom Value' row, and auto-select it.", 
     "Not Run", "Critical"),
    ("F-COR-06", "Corrections", "Custom Clear", 
     "Click 'X' (Clear) in the Autocomplete field", 
     "Custom suggestion row should disappear, and 'Accept' button should revert to disabled.", 
     "Not Run", "High"),

    # ── SUBMISSION WORKFLOWS ──────────────────────────────────────────────
    ("F-SUB-01", "Submission", "Accept Flow", 
     "Select suggestion -> Click Accept -> Confirm in dialog", 
     "PUT request to /approve-suggestion with correct payload. On success: Snackbar shows, field expands to 'Accepted' status.", 
     "Not Run", "Critical"),
    ("F-SUB-02", "Submission", "Send to L0 Flow", 
     "Click 'Send to L0' on a pending field -> Confirm in dialog", 
     "PUT request to /reject-record. On success: Snackbar shows, field collapses with 'L0 Data' status.", 
     "Not Run", "Critical"),
    ("F-SUB-03", "Submission", "Draft Submission", 
     "Click 'Submit Draft Record' -> Fill dynamic form -> Submit", 
     "POST request to /draft-records/<field>. On success: Snackbar shows, and 'Draft Record' row appears in table.", 
     "Not Run", "Critical"),
    ("F-SUB-04", "Submission", "Draft Validation", 
     "Enter decimal or negative value in an 'integer' type draft field -> Submit", 
     "Submit should be blocked by frontend validation. Snackbar should show error message.", 
     "Not Run", "High"),
    ("F-SUB-05", "Submission", "Draft Field Loading", 
     "Open Draft Dialog", 
     "Fields should load dynamically from /draft-records/fields API. History values should pre-fill where applicable.", 
     "Not Run", "Medium"),

    # ── FEEDBACK & FEEDBACK ───────────────────────────────────────────────
    ("F-FB-01", "Feedback", "Snackbar", 
     "Trigger a successful action", 
     "Snackbar with 'success' variant should appear at top center and auto-close after 3 seconds.", 
     "Not Run", "Medium"),
    ("F-FB-02", "Feedback", "Confirmation Lock", 
     "Click Submit/Confirm and then try to click Cancel while submitting", 
     "Buttons should be disabled/locked during the API flight to prevent duplicate submissions.", 
     "Not Run", "High"),
]

# Create Workbook
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Functional Interactable QA"
ws.sheet_view.showGridLines = False

# Header
ws.merge_cells("A1:G1")
header = ws.cell(row=1, column=1, value="Core Functional Interactions — QA Report")
header.fill = fill(C["navy"]); header.font = font(bold=True, size=14, color=C["white"]); header.alignment = align("center")

# Table Headers
cols = ["TC ID", "Functional Area", "Component", "Interaction / Step", "Expected Behavioral Result", "Status", "Priority"]
for i, h in enumerate(cols, 1):
    cell = ws.cell(row=2, column=i, value=h)
    cell.fill = fill(C["navy"]); cell.font = font(bold=True, color=C["white"]); cell.alignment = align("center")

# Column Widths
widths = [10, 20, 25, 45, 60, 15, 12]
for i, w in enumerate(widths, 1):
    ws.column_dimensions[get_column_letter(i)].width = w

# Data
for r, (tc, area, comp, inter, exp, stat, pri) in enumerate(FUNCTIONAL_TESTS, 3):
    bg = C["white"] if r % 2 == 0 else C["gray_bg"]
    ws.cell(row=r, column=1, value=tc).fill = fill(bg)
    ws.cell(row=r, column=2, value=area).fill = fill(bg)
    ws.cell(row=r, column=3, value=comp).fill = fill(bg)
    ws.cell(row=r, column=4, value=inter,).fill = fill(bg)
    ws.cell(row=r, column=5, value=exp).fill = fill(bg)
    
    # Status Cell
    s_cell = ws.cell(row=r, column=6, value=stat)
    s_cell.fill = fill(C["notrun_bg"]); s_cell.font = font(bold=True, color=C["notrun_fg"]); s_cell.alignment = align("center")
    
    # Priority Cell
    p_cell = ws.cell(row=r, column=7, value=pri)
    p_cell.fill = fill(bg); p_cell.alignment = align("center")
    p_color = {"Critical": "FF991B1B", "High": "FF9A3412", "Medium": "FF854D0E"}.get(pri, "FF000000")
    p_cell.font = font(bold=True, color=p_color)

    # Borders
    for c in range(1, 8):
        ws.cell(row=r, column=c).border = bdr()
        ws.cell(row=r, column=c).alignment = align(wrap=True)

# Save
OUTPUT_PATH = r"d:\DH\Data-Hygiene\QA_Functional_Interactions_Report.xlsx"
wb.save(OUTPUT_PATH)
print(f"Report saved to: {OUTPUT_PATH}")
